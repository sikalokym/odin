from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd

from src.database.db_operations import DBOperations


def get_sheet(ws, entities_ids_dict, pnos_ids, title, time, country):

    time_str = str(time) + '0'
    date = pd.to_datetime(time_str, format='%Y%U%w').date()
    conditions = [f"CountryCode = '{country}'", f"CAST(ChangeDate AS DATE) < '{date}'"]
    or_conditions = []
    for table, ids in entities_ids_dict.items():
        table_name = DBOperations.instance.config.get('TABLES', table)
        if len(ids) == 0:
            continue
        elif len(ids) == 1:
            or_conditions.append(f"ChangeTable = '{table_name}' AND ChangeCode = '{ids[0]}'")
        else:
            or_conditions.append(f"ChangeTable = '{table_name}' AND ChangeCode IN {tuple(ids)}")
    if len(or_conditions) != 0:
        conditions.append('(' + ' OR '.join(or_conditions) + ')')
    
    df_change_log = DBOperations.instance.get_table_df(DBOperations.instance.config.get('DQ', 'CL'), conditions=conditions)

    pno_tables = ['COL', 'UPH', 'OPT', 'PKG', 'FEAT', 'CFEAT']
    conditions = []
    if len(pnos_ids) == 1:
        conditions.append(f"ChangeCode = '{pnos_ids[0]}'")
    else:
        conditions.append(f"ChangeCode in {tuple(pnos_ids)}")
    or_conditions = [ f"ChangeTable = '{DBOperations.instance.config.get('AUTH', table)}'" for table in pno_tables]
    or_cond = '(' + ' OR '.join(or_conditions) + ')' if len(or_conditions) != 0 else ''
    conditions.append(f"CAST(ChangeDate AS DATE) < '{date}'")
    conditions.append(or_cond)

    df_pno_change_log = DBOperations.instance.get_table_df(DBOperations.instance.config.get('DQ', 'CL'), conditions=conditions)

    df_change_log = pd.concat([df_change_log, df_pno_change_log])
    
    # relations_tables = ['PNO_Custom', 'PNOColorCustom', 'PNOOptionsCustom', 'PNOUpholsteryCustom', 'PNOPackageCustom']

    # Write the headers
    ws['A1'] = f'{title} - Änderungen'
    ws.append(['Date', 'Entity', 'Change'])
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:C1')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.freeze_panes = ws['A2']

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 47
    ws.column_dimensions['C'].width = 75

    ws.row_dimensions[1].height = 45
    ws.row_dimensions[2].height = 35

    # Formatting of first row
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')
    ws['A1'].fill = fill
    ws['B1'].fill = fill
    ws['C1'].fill = fill
    
    # Formatting of second row
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center',wrap_text=True)
    ws['A2'].font = Font(size=10, bold=True)
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    ws['B2'].font = Font(size=10, bold=True)
    ws['C2'].alignment = Alignment(horizontal='left', vertical='center')
    ws['C2'].font = Font(size=10, bold=True)

    if df_change_log.empty:
        return None
    for change in df_change_log.itertuples():
       ws.append([change.ChangeDate.strftime("%Y-%m-%d"), f'{change.ChangeTable}', f'{change.ChangeField} hat sich geändert: {change.ChangeFrom} -> {change.ChangeTo}'])
