from openpyxl.styles import PatternFill, Border, Alignment, Border, Side, Font
import pandas as pd
import numpy as np

from src.utils.db_utils import filter_df_by_timestamp, format_float_string
from src.database.db_operations import DBOperations

# @author Hassan Wahba

all_border = Border(top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'),
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'))

fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')

def get_sheet(ws, sales_versions, title, time, cell_values, rule_texts, config):
    """
    Fetches options data and inserts it into the specified worksheet.

    Args:
        ws (Worksheet): The worksheet to insert the data into.
        sales_versions (DataFrame): The sales versions to fetch options data for.
        title (str): The title of the sheet.

    Returns:
        None
    """
    df_upholstery = fetch_upholstery_data(sales_versions.copy(), time, rule_texts)
    df_colors = fetch_color_data(sales_versions.copy(), time, rule_texts)

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 65

    insert_table(ws, sales_versions, title + ' - ' + config['UPHOLSTERY_COLORS_SHEET']['UPHOLSTERY'], df_upholstery, cell_values, config['DEFAULT']['TAX'].replace('\\n', '\n'), config['UPHOLSTERY_COLORS_SHEET']['NOTES'])
    mid_row = ws.max_row + 1
    ws.append([''])
    insert_table(ws, sales_versions, title + ' - ' + config['UPHOLSTERY_COLORS_SHEET']['COLOR'], df_colors, cell_values, config['DEFAULT']['TAX'].replace('\\n', '\n'), config['UPHOLSTERY_COLORS_SHEET']['NOTES'])
    
    for col in range(3, ws.max_column):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
    ws.column_dimensions[ws.cell(row=1, column=ws.max_column).column_letter].width = 40

    ws.column_dimensions[ws.cell(row=1, column=ws.max_column-1).column_letter].width = 1
    for cell in ws[ws.cell(row=1, column=ws.max_column-1).column_letter]:
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        if cell.row == mid_row:
            cell.border = Border(top=None, bottom=None, left=None, right=None)
        else:
            cell.border = Border(top=None, bottom=None, left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'))

    ws.sheet_view.showGridLines = False

def insert_table(ws, sales_versions, title, df_res, cell_values, tax, notes):

    insert_title(ws, sales_versions, title, tax, notes)
    old_custom_category = -1
    catgory_rows = []
    first_row = ws.max_row + 1
    row_height_factor = 15
    for _, row in df_res.iterrows():
        if 'CustomCategory' in df_res.columns and row['CustomCategory'] != old_custom_category:
            row_height_factor = 25
            add_custom_category(ws, row['CustomCategory'])
            old_custom_category = row['CustomCategory']
            catgory_rows.append(ws.max_row)
        
        svs = []
        for sv in sales_versions['SalesVersion']:
            if sv in df_res.columns:
                if pd.isna(row[sv]):
                    svs.append('')
                    continue
                parts = row[sv].split('\n')
                if len(parts) > 1:
                    svs.append(cell_values.get(parts[0], parts[0]) + '\n' + parts[1])
                else:
                    svs.append(cell_values.get(parts[0], parts[0]))
            else:
                svs.append('')
        if row['Price'] == 'Pack Only'or row['Price'] == 'Serie':
            ws.append([row['Code'], row['CustomName'], row['Price']] + svs + ['', row['Rules']])
            row_height = int(len(row['CustomName']) / 65) * 15 + row_height_factor
            ws.row_dimensions[ws.max_row].height = row_height
            ws.append([''])
        else:
            prices = row['Price'].split('/')
            if len(prices) > 1:
                ws.append([row['Code'], row['CustomName'], prices[0]] + svs + ['', row['Rules']])
                row_height = int(len(row['CustomName']) / 65) * 15 + row_height_factor
                ws.row_dimensions[ws.max_row].height = row_height
                ws.cell(row=ws.max_row, column=3).alignment = Alignment(horizontal='center', vertical='bottom')
                ws.append(['', '', prices[1]] + svs) 
                ws.cell(row=ws.max_row, column=3).alignment = Alignment(horizontal='center', vertical='top')
        format_row(ws, ws.max_row-1, ws.max_row)
    
    format_first_two_columns(ws, first_row, ws.max_row, catgory_rows)

def format_first_two_columns(ws, row1, row2, catgory_rows):
    # for each row in the first column, merge the cells in the first two columns if the cell's value is the same as the one above
    max_col_letter = ws.cell(row=row1, column=ws.max_column).column_letter
    entry_rows = []
    while row1 <= row2:
        if row1 not in catgory_rows:
            entry_rows.append(row1)
            row1 += 1
        row1 += 1

    curr_row = entry_rows[0]
    curr_end_row = entry_rows[0]
    for row in entry_rows[1:]:
        if ws.cell(row=row, column=1).value == ws.cell(row=curr_row, column=1).value and ws.cell(row=curr_row, column=1).value:
            curr_end_row = row
        elif curr_row != curr_end_row:
            for sub_row in range(curr_row, curr_end_row+1, 2):
                ws.unmerge_cells(f'A{sub_row}:A{sub_row+1}')
                ws.unmerge_cells(f'B{sub_row}:B{sub_row+1}')
                ws.unmerge_cells(f'{max_col_letter}{sub_row}:{max_col_letter}{sub_row+1}')

            ws.merge_cells(f'A{curr_row}:A{curr_end_row+1}')
            ws.merge_cells(f'B{curr_row}:B{curr_end_row+1}')
            ws.merge_cells(f'{max_col_letter}{curr_row}:{max_col_letter}{curr_end_row+1}')
            curr_row = row
            curr_end_row = row
        else:
            curr_row = row
            curr_end_row = row

def add_custom_category(ws, custom_category):
    ws.append(['', custom_category])
    # add border around the second cell of the row
    ws.cell(row=ws.max_row, column=2).border = Border(top=None, bottom=None, left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'))
    for col in range(1, ws.max_column+1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if col == ws.max_column:
            cell.border = Border(top=None, bottom=None, left=None, right=Side(style='thin', color='000000'))
        elif col == 3:
            cell.border = Border(top=None, bottom=Side(style='thin', color='000000'), left=None, right=None)

def format_row(ws, row1, row2):
    for col in range(1, ws.max_column+1):
        if col != 3:
            cell = ws.cell(row=row1, column=col)
            alig = 'center' if col > 2 else 'left'
            cell.alignment = Alignment(horizontal=alig, vertical='center', wrap_text=True)
            cell.border = all_border
            cell.font = Font(name='Arial', size=10)
            ws.merge_cells(f'{ws.cell(row=row1, column=col).column_letter}{row1}:{ws.cell(row=row2, column=col).column_letter}{row2}')
        else:
            ws.cell(row=row1, column=col).alignment = Alignment(horizontal='center', vertical='bottom')
            ws.cell(row=row1, column=col).font = Font(name='Arial', size=10, bold=True)
            ws.cell(row=row2, column=col).alignment = Alignment(horizontal='center', vertical='top')
            ws.cell(row=row2, column=col).border = Border(top=None, bottom=Side(style='thin', color='000000'), left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'))

def insert_title(ws, sales_versions, title, tax, notes):

    # Content of Row 1 & 2
    ws.append([title, '', tax] + [f'{row["SalesVersionName"]}\nSV {row["SalesVersion"]}' for _, row in sales_versions.iterrows()] + ['', notes])
    insert_row_index = ws.max_row
    ws.merge_cells(f'A{insert_row_index}:B{insert_row_index}')
    
    # fill color
    for col in range(1, ws.max_column+1):
        cell = ws.cell(row=insert_row_index, column=col)
        cell.fill = fill
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

def fetch_color_data(sales_versions, time, rule_texts):
    pno_ids = sales_versions.ID.unique().tolist()
    conditions = []
    if len(pno_ids) == 1:
        conditions.append(f"PNOID = '{pno_ids[0]}'")
    else:
        conditions.append(f"PNOID in {tuple(pno_ids)}")
    df_pno_color = DBOperations.instance.get_table_df(DBOperations.instance.config.get('AUTH', 'COL'), columns=['ID', 'PNOID', 'Code', 'RuleName', 'StartDate', 'EndDate'], conditions=conditions)
    df_pno_color = filter_df_by_timestamp(df_pno_color, time)
    rel_codes = df_pno_color.ID.unique().tolist()
    pno_color_price_conditions = []
    if len(rel_codes) == 1:
        pno_color_price_conditions.append(f"RelationID = '{rel_codes[0]}'")
    else:
        pno_color_price_conditions.append(f"RelationID in {tuple(rel_codes)}")
    df_pno_color_price = DBOperations.instance.get_table_df(DBOperations.instance.config.get('RELATIONS', 'COL_Custom'), columns=['RelationID', 'Price', 'PriceBeforeTax', 'CustomName'], conditions=pno_color_price_conditions)
    df_pno_color_price = df_pno_color_price.drop_duplicates()

    sales_versions = sales_versions.rename(columns={'ID': 'TmpCode'})
    df_pno_color = df_pno_color.merge(sales_versions[['TmpCode', 'SalesVersion', 'SalesVersionName']], left_on='PNOID', right_on='TmpCode', how='left')
    df_pno_color = df_pno_color.drop(columns='TmpCode')
    df_pno_color_with_sv = df_pno_color[df_pno_color['SalesVersion'].notna()]

    # Replace ID with Price from df_pno_color_price
    df_pno_color_with_price = df_pno_color_with_sv.merge(df_pno_color_price, left_on='ID', right_on='RelationID', how='left')

    # format prices to float using format_float_string
    df_pno_color_with_price['Price'] = df_pno_color_with_price['Price'].apply(format_float_string)
    df_pno_color_with_price['PriceBeforeTax'] = df_pno_color_with_price['PriceBeforeTax'].apply(format_float_string)

    # Concatenate Price and PriceBeforeTax
    df_pno_color_with_price['Price'] = df_pno_color_with_price.apply(lambda x: f"{x['Price']}/{x['PriceBeforeTax']}", axis=1)
    df_pno_color_with_price = df_pno_color_with_price.drop_duplicates()
    
    # Group by 'Code' and 'Price', aggregate 'CustomName' and 'CustomCategory' columns
    aggregated = df_pno_color_with_price.groupby(['Code', 'Price']).agg({
        'CustomName': lambda x: ';\n'.join(sorted([y for y in x.dropna().unique() if y]))
    }).reset_index()
    df_pno_color_with_price = df_pno_color_with_price.drop(columns=['CustomName'])
    
    # Merge the aggregated columns back to the original dataframe
    df_merged = df_pno_color_with_price.merge(aggregated, on=['Code', 'Price'])

    # group by Code, Price, RuleName, SalesVersion and SalesVersionName and aggregate the CustomName column by concatinating the values with a newline separator if they differ and not null nan or empty and the CustomCategory column by concatinating the values with a comma separator if they differ
    df_pno_color_with_price = df_merged.groupby(['Code', 'Price', 'RuleName', 'SalesVersion', 'SalesVersionName']).agg({'CustomName': 'first'}).reset_index()
    
    # Create the pivot table
    pivot_df = df_pno_color_with_price.pivot_table(index=['Code', 'Price'], columns='SalesVersion', values='RuleName', aggfunc='first')

    # Drop the now unneeded columns and duplicates
    df_pno_color_with_price = df_pno_color_with_price.drop(['RuleName', 'SalesVersion', 'SalesVersionName'], axis=1).drop_duplicates()

    # Join the pivoted DataFrame with the original one. sort after code ascending
    df_result = df_pno_color_with_price.join(pivot_df, on=['Code', 'Price']).sort_values(by='Code')

    # get the df with the rules for the upholstery
    opt_codes = df_result['Code'].unique().tolist()
    rules_conditions = conditions.copy() + ["RuleCode LIKE 'U%'"]
    if len(opt_codes) == 1:
        rules_conditions.append(f"ItemCode = '{opt_codes[0]}'")
    else:
        rules_conditions.append(f"ItemCode in {tuple(opt_codes)}")
    df_rules = DBOperations.instance.get_table_df(DBOperations.instance.config.get('DEPENDENCIES', 'UFO'), columns=['RuleCode', 'ItemCode', 'FeatureCode'], conditions=rules_conditions)
    df_rules = df_rules.drop_duplicates()

    df_result['Rules'] = ''
    if df_rules.empty:
        return df_result

    # strip the whitespace from the ItemCode and FeatureCode columns
    df_rules['ItemCode'] = df_rules['ItemCode'].str.strip()
    df_rules['FeatureCode'] = df_rules['FeatureCode'].str.strip()
    
    for rule, group in df_rules.groupby('RuleCode'):
        group = group.groupby('ItemCode').agg({'FeatureCode': lambda x: rule_texts.get(rule.lower(), rule) + ' ' + ', '.join(list(x))}).reset_index()
        group = group.rename(columns={'FeatureCode': rule})
        df_result = pd.merge(df_result, group, left_on='Code', right_on='ItemCode', how='left')
        df_result = df_result.drop(columns=['ItemCode'])
        # accumulate the rules in the Rules column with a new line separator
        df_result['Rules'] = df_result.apply(lambda row: row['Rules'] + '\n' + row[rule] if pd.notnull(row[rule]) else row['Rules'], axis=1)
        df_result = df_result.drop(columns=[rule])
    
    # remove the first new line separator
    df_result['Rules'] = df_result['Rules'].apply(lambda x: x[1:] if x.startswith('\n') else x)
    
    df_result = df_result.fillna('')
    return df_result

def fetch_upholstery_data(sales_versions, time, rule_texts):
    pno_ids = sales_versions.ID.unique().tolist()
    conditions = []
    if len(pno_ids) == 1:
        conditions.append(f"PNOID = '{pno_ids[0]}'")
    else:
        conditions.append(f"PNOID in {tuple(pno_ids)}")
    df_pno_upholstery = DBOperations.instance.get_table_df(DBOperations.instance.config.get('AUTH', 'UPH'), columns=['ID', 'PNOID', 'Code', 'RuleName', 'StartDate', 'EndDate'], conditions=conditions)
    df_pno_upholstery = filter_df_by_timestamp(df_pno_upholstery, time)
    rel_codes = df_pno_upholstery.ID.unique().tolist()
    pno_upholstery_price_conditions = []
    if len(rel_codes) == 1:
        pno_upholstery_price_conditions.append(f"RelationID = '{rel_codes[0]}'")
    else:
        pno_upholstery_price_conditions.append(f"RelationID in {tuple(rel_codes)}")
    df_pno_upholstery_price = DBOperations.instance.get_table_df(DBOperations.instance.config.get('RELATIONS', 'UPH_Custom'), columns=['RelationID', 'Price', 'PriceBeforeTax', 'CustomName', 'CustomCategory'], conditions=pno_upholstery_price_conditions)
    df_pno_upholstery_price = df_pno_upholstery_price.drop_duplicates()
    
    # Load features that
    feat_conditions = conditions.copy() + ["Code LIKE 'NC%'"]
    df_pno_features = DBOperations.instance.get_table_df(DBOperations.instance.config.get('AUTH', 'FEAT'), columns=['PNOID', 'Reference as Code', 'CustomName as DecorInlay'], conditions=feat_conditions)
    df_pno_features['Code'] = df_pno_features['Code'].apply(lambda x: x.strip().split('(u)')[0])
    
    df_pno_features = df_pno_features[df_pno_features['Code'].isin(df_pno_upholstery['Code'].unique().tolist())]
    
    # Remove Options that have a subset of the relatable features
    for ref in df_pno_features['Code'].unique():
        local_group = df_pno_features[df_pno_features['Code'] == ref].groupby(['PNOID']).agg({'Code': lambda x: len(x)})
        max_num_feats = local_group['Code'].max()
        # Get sales versions with the maximum number of features
        max_sales_versions = local_group[local_group['Code'] != max_num_feats].index.tolist()
        if len(max_sales_versions):
            df_pno_features = df_pno_features[~((df_pno_features['Code'] == ref) & (df_pno_features['PNOID'].isin(max_sales_versions)))]
    
    df_pno_upholstery = df_pno_upholstery.merge(df_pno_features, on=['PNOID', 'Code'], how='left')
    
    # Condition to check not NaN, not None, and not empty
    condition = df_pno_upholstery['DecorInlay'].notna() & (df_pno_upholstery['DecorInlay'] != '')

    # Apply the condition and update the RuleName column
    df_pno_upholstery['RuleName'] = np.where(condition,
                                            df_pno_upholstery['RuleName'] + '\n' + df_pno_upholstery['DecorInlay'],
                                            df_pno_upholstery['RuleName'])
    df_pno_upholstery = df_pno_upholstery.drop(columns='DecorInlay')
    
    sales_versions = sales_versions.rename(columns={'ID': 'TmpCode'})
    df_pno_upholstery = df_pno_upholstery.merge(sales_versions[['TmpCode', 'SalesVersion', 'SalesVersionName']], left_on='PNOID', right_on='TmpCode', how='left')
    df_pno_upholstery = df_pno_upholstery.drop(columns='TmpCode')
    df_pno_upholstery_with_sv = df_pno_upholstery[df_pno_upholstery['SalesVersion'].notna()].drop_duplicates()

    # Replace ID with Price from df_pno_upholstery_price
    df_pno_upholstery_with_price = df_pno_upholstery_with_sv.merge(df_pno_upholstery_price, left_on='ID', right_on='RelationID', how='left')

    # format prices to float using format_float_string
    df_pno_upholstery_with_price['Price'] = df_pno_upholstery_with_price['Price'].apply(format_float_string)
    df_pno_upholstery_with_price['PriceBeforeTax'] = df_pno_upholstery_with_price['PriceBeforeTax'].apply(format_float_string)

    # Concatenate Price and PriceBeforeTax
    df_pno_upholstery_with_price['Price'] = df_pno_upholstery_with_price.apply(lambda x: f"{x['Price']}/{x['PriceBeforeTax']}", axis=1)
    
    # Group by 'Code' and 'Price', aggregate 'CustomName' and 'CustomCategory' columns
    aggregated = df_pno_upholstery_with_price.groupby(['Code', 'Price']).agg({
        'CustomName': lambda x: ';\n'.join(sorted([y for y in x.dropna().unique() if y])),
        'CustomCategory': lambda x: ';\n'.join(sorted([y for y in x.dropna().unique() if y]))
    }).reset_index()
    
    df_pno_upholstery_with_price = df_pno_upholstery_with_price.drop(columns=['CustomName', 'CustomCategory'])
    
    # Merge the aggregated columns back to the original dataframe
    df_merged = df_pno_upholstery_with_price.merge(aggregated, on=['Code', 'Price'])

    # Group by Code, Price, RuleName, SalesVersion and SalesVersionName and aggregate the CustomName column by concatinating the values with a newline separator if they differ and not null nan or empty and the CustomCategory column by concatinating the values with a comma separator if they differ
    df_pno_upholstery_with_price = df_merged.groupby(['Code', 'Price', 'RuleName', 'SalesVersion', 'SalesVersionName']).agg({'CustomName': 'first', 'CustomCategory': 'first'}).reset_index()
        
    # Create the pivot table
    pivot_df = df_pno_upholstery_with_price.pivot_table(index=['Code', 'Price'], columns='SalesVersion', values='RuleName', aggfunc='first')

    # Drop the now unneeded columns and duplicates
    df_pno_upholstery_with_price = df_pno_upholstery_with_price.drop(['RuleName', 'SalesVersion', 'SalesVersionName'], axis=1).drop_duplicates()

    # Join the pivoted DataFrame with the original one. sort after code ascending
    df_result = df_pno_upholstery_with_price.join(pivot_df, on=['Code', 'Price']).sort_values(by='Code')

    # get the df with the rules for the upholstery
    opt_codes = df_result['Code'].unique().tolist()
    rules_conditions = conditions.copy() + ["RuleCode LIKE 'U%'"]
    if len(opt_codes) == 1:
        rules_conditions.append(f"ItemCode = '{opt_codes[0]}'")
    else:
        rules_conditions.append(f"ItemCode in {tuple(opt_codes)}")
    df_rules = DBOperations.instance.get_table_df(DBOperations.instance.config.get('DEPENDENCIES', 'UFO'), columns=['RuleCode', 'ItemCode', 'FeatureCode'], conditions=rules_conditions)
    df_rules = df_rules.drop_duplicates()

    df_result['Rules'] = ''
    if df_rules.empty:
        return df_result.sort_values('CustomCategory')
    
    # strip the whitespace from the ItemCode and FeatureCode columns
    df_rules['ItemCode'] = df_rules['ItemCode'].str.strip()
    df_rules['FeatureCode'] = df_rules['FeatureCode'].str.strip()
    
    for rule, group in df_rules.groupby('RuleCode'):
        group = group.groupby('ItemCode').agg({'FeatureCode': lambda x: rule_texts.get(rule.lower(), rule) + ' ' + ', '.join(list(x))}).reset_index()
        group = group.rename(columns={'FeatureCode': rule})
        df_result = pd.merge(df_result, group, left_on='Code', right_on='ItemCode', how='left')
        df_result = df_result.drop(columns=['ItemCode'])
        # accumulate the rules in the Rules column with a new line separator
        df_result['Rules'] = df_result.apply(lambda row: row['Rules'] + '\n' + row[rule] if pd.notnull(row[rule]) else row['Rules'], axis=1)
        df_result = df_result.drop(columns=[rule])
    
    # remove the first new line separator
    df_result['Rules'] = df_result['Rules'].apply(lambda x: x[1:] if x.startswith('\n') else x)

    df_result = df_result.sort_values('CustomCategory')
    
    df_result = df_result.fillna('')
    return df_result
