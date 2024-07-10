from openpyxl.styles import PatternFill, Border, Alignment, Side

import src.export.variant_binder.options_sheet as options_sheet

#General formating of border lines & cell colours in excel spreadsheet
all_border = Border(top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'),
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'))

white_border = Border(right=Side(style='thin', color="FFFFFF"))
fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

def get_sheet(ws, sales_versions, title, df_res, cell_values, config):
    """
    Fetches options data and inserts it into the specified worksheet.

    Args:
        ws (Worksheet): The worksheet to insert the data into.
        sales_versions (DataFrame): The sales versions to fetch options data for.
        title (str): The title of the sheet.

    Returns:
        None 
    """
    ws.append([''])
    ws.append([''])
    
    # insert data into the sheet ab row 3
    for _, row in df_res.iterrows():
        svs = [cell_values.get(row[sv], row[sv]) if sv in df_res.columns else '' for sv in sales_versions['SalesVersion']]
        if row['Price'] == config['TIRES_SHEET']['GRAYED_OUT_PRICE']or row['Price'] == config['DEFAULT']['SERIES']:
            ws.append([row['Code'], row['CustomName'], row['Price']] + svs + ['', row['CustomCategory'], row['Rules']])
            row_height = int(len(row['CustomName']) / 65) * 15 + 15
            ws.row_dimensions[ws.max_row].height = row_height
            ws.append([''])
        else:
            prices = row['Price'].split('/')
            if len(prices) > 1:
                ws.append([row['Code'], row['CustomName'], prices[0]] + svs + ['', row['CustomCategory'], row['Rules']])
                ws.cell(row=ws.max_row, column=3).alignment = Alignment(horizontal='center', vertical='bottom')
                row_height = int(len(row['CustomName']) / 65) * 15 + 15
                ws.row_dimensions[ws.max_row].height = row_height
                ws.append(['', '', prices[1]] + svs) 
                ws.cell(row=ws.max_row, column=3).alignment = Alignment(horizontal='center', vertical='top')

    options_sheet.prepare_sheet(ws, sales_versions, f'{title} - ' + config['TIRES_SHEET']['TITLE'], config)

    # remove before last column with content
    ws.delete_cols(ws.max_column - 1)
