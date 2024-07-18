from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import pandas as pd

from src.database.db_operations import DBOperations


all_border = Border(top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'),
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'))

fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')

def get_sheet(ws, sales_versions, title, config):
    """
    Fetches options data and inserts it into the specified worksheet.

    Args:
        ws (Worksheet): The worksheet to insert the data into.
        sales_versions (DataFrame): The sales versions to fetch options data for.
        title (str): The title of the sheet.

    Returns:
        None
    """
    df_sales_versions = fetch_sales_version_data(sales_versions)

    # Write the headers
    ws['A1'] = f'{title} - {config["title"]}'
    ws.append([config['code_column'], config['description_column']])
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:B1')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.freeze_panes = ws['A2']

    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 150

    ws.row_dimensions[1].height = 45
    ws.row_dimensions[2].height = 35

    # Formatting of first row
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws['A1'].fill = fill
    ws['B1'].fill = fill
    
    # Formatting of second row
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    ws['A2'].font = Font(size=10, bold=True)
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    ws['B2'].font = Font(size=10, bold=True)

    if df_sales_versions.empty:
        return
    svn_with_features = df_sales_versions.SalesVersionName.unique().tolist()
    last_svn = svn_with_features[-1] if len(svn_with_features) > 0 else None
    last_sv = df_sales_versions[df_sales_versions.SalesVersionName == last_svn].SalesVersion.unique()[0]
    svns_without_features = sales_versions[~sales_versions.SalesVersionName.isin(svn_with_features)]
    combined_sv = ' / '.join([last_sv] + svns_without_features.SalesVersion.unique().tolist())
    combuned_svn = ' / '.join([last_svn] + svns_without_features.SalesVersionName.unique().tolist())
    
    prev_svn = []
    for (sv, svn), group in df_sales_versions.groupby(['SalesVersion', 'SalesVersionName'], observed=False):
        sv = sv if sv != last_sv else combined_sv
        svn = svn if svn != last_svn else combuned_svn
        group = group.drop(columns=['SalesVersion', 'SalesVersionName'])
        df_options = group.sort_values(by=['CustomCategory', 'CustomName'], ascending=True)

        # Write the data to the worksheet
        prev_sv_text = prev_svn[-1] if len(prev_svn) > 0 else None
        svn_with_name = f"{svn} {config['category_dependency'].replace('{prev_sv}', prev_sv_text)})" if len(prev_svn) > 0 else svn
        if svn not in prev_svn:
            prev_svn.append(svn)
        ws.append([''])
        ws.append([sv, svn_with_name])
        # format the row in gray
        for cell in ws[len(ws["A"])]:
            cell.fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
        prev_custom_category = None
        for _, row in df_options.iterrows():
            current_custom_category = row['CustomCategory']
            if prev_custom_category != current_custom_category:
                ws.append([''])
                ws.append(['', row['CustomCategory']])
                # format the text in bold
                for cell in ws[len(ws["A"])]:
                    cell.font = Font(bold=True)
                prev_custom_category = current_custom_category
            ws.append([row['Code'], row['CustomName']])
    
    for col in range(1, 3):
        cell = ws.cell(row=ws.max_row, column=col)
        if col == 1:
            cell.border = Border(bottom=Side(style='medium'),
                                right=Side(style='thin'))
        else:
            cell.border = Border(bottom=Side(style='medium'),
                                right=Side(style='medium'))
            
    for row in range(1, ws.max_row):
        cell = ws.cell(row=row, column=1)
        cell.border = Border(bottom=Side(style='thin'),
                            right=Side(style='thin'))
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        cell = ws.cell(row=row, column=2)
        cell.border = Border(bottom=Side(style='thin'),
                            right=Side(style='medium'))
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
            
def fetch_sales_version_data(df_sales_versions):
    pno_ids = df_sales_versions.ID.unique().tolist()
    # filter where code starts with X
    conditions = []
    if len(pno_ids) == 1:
        conditions.append(f"PNOID = '{pno_ids[0]}'")
    else:
        conditions.append(f"PNOID in {tuple(pno_ids)}")
    features_conditions = conditions.copy() + [f"Code not like 'X%'"]
    df_pno_features_all = DBOperations.instance.get_table_df(DBOperations.instance.config.get('AUTH', 'FEAT'), columns=['PNOID', 'Code', 'Reference', 'RuleName', 'CustomName', 'CustomCategory'], conditions=features_conditions)
    df_pno_features = df_pno_features_all[df_pno_features_all['RuleName'].str.strip() == 'S'].drop(columns=['RuleName'])
    df_pno_custom_features = DBOperations.instance.get_table_df(DBOperations.instance.config.get('AUTH', 'CFEAT'), columns=['PNOID', 'Code', 'CustomName', 'CustomCategory'], conditions=conditions)
    
    refs = df_pno_features['Reference'].apply(lambda x: x.strip().zfill(6)).unique().tolist()
    ref_conds = conditions.copy()
    if len(refs) == 1:
        ref_conds.append(f"RuleCode = '{refs[0]}'")
    else:
        ref_conds.append(f"RuleCode in {tuple(refs)}")
    df_pno_pkgs = DBOperations.instance.get_table_df(DBOperations.instance.config.get('AUTH', 'PKG'), conditions=conditions)
    if not df_pno_pkgs.empty:
        cus_conds = ['Price = 0']
        ids = df_pno_pkgs.ID.unique().tolist()
        if len(ids) == 1:
            cus_conds.append(f"RelationID = '{ids[0]}'")
        else:
            cus_conds.append(f"RelationID in {tuple(ids)}")
        df_custom_pkgs = DBOperations.instance.get_table_df(DBOperations.instance.config.get('RELATIONS', 'PKG_Custom'), columns=['RelationID'], conditions=cus_conds)
        if not df_custom_pkgs.empty:
            rel_ids = df_custom_pkgs.RelationID.unique().tolist()
            df_pkg_full = df_pno_pkgs[df_pno_pkgs.ID.isin(rel_ids)]
            for code in df_pkg_full.RuleCode.unique():
                feats = df_pno_features[df_pno_features['Reference'].str.strip().apply(lambda x: x.zfill(6)) == code]
                if feats.empty:
                    new_rows = df_pkg_full[df_pkg_full['RuleCode'] == code][['PNOID', 'RuleCode']]
                    new_rows['Reference'] = new_rows['RuleCode'].apply(lambda x: x.lstrip('0'))
                    df_merged_rows = new_rows[['PNOID', 'Reference']].merge(df_pno_features_all[['PNOID', 'Code', 'Reference', 'CustomName', 'CustomCategory']], on=['PNOID', 'Reference'], how='inner')
                    df_pno_features = pd.concat([df_pno_features, df_merged_rows], ignore_index=True)
                feat_pnos = feats.PNOID.unique().tolist()
                should_pnos = df_pkg_full[(df_pkg_full['RuleCode'] == code) & (~df_pkg_full['PNOID'].isin(feat_pnos))].PNOID.unique().tolist()
                if not should_pnos:
                    continue
                for pno in should_pnos:
                    temple_row = feats.iloc[:1].copy()
                    temple_row['PNOID'] = pno
                    df_pno_features = pd.concat([df_pno_features, temple_row], ignore_index=True)
    
    # Code should have (reference) appendend to it, if Reference is not null and not empty
    df_pno_features['Code'] = df_pno_features['Code'].str.strip()
    df_pno_features = df_pno_features.groupby(['Code', 'PNOID']).agg({
        'Reference': lambda x: ', '.join([y for y in set(x) if y != '' and y is not None]),
        'CustomName': 'first',
        'CustomCategory': 'first'}).reset_index()
    df_pno_features['Code'] = df_pno_features.apply(lambda x: f"{x['Code']} ({x['Reference']})" if x['Reference'] else x['Code'], axis=1)
    df_pno_features = df_pno_features.drop(columns=['Reference'])
    df_pno_features = pd.concat([df_pno_features, df_pno_custom_features])
    sv_correct_order = df_sales_versions.SalesVersion.unique().tolist()

    df_pno_features_sv = df_pno_features.merge(df_sales_versions[['ID', 'SalesVersion', 'SalesVersionName']], left_on='PNOID', right_on='ID', how='inner')
    df_pno_features_sv['SalesVersion'] = pd.Categorical(df_pno_features_sv['SalesVersion'], sv_correct_order)
    df_pno_features_sv = df_pno_features_sv.sort_values(by='SalesVersion')

    df_pno_features_sv = df_pno_features_sv.drop_duplicates('Code', keep='first').drop(columns=['ID', 'PNOID'])
    
    df_pno_features_sv = df_pno_features_sv.dropna(subset=['CustomName'])
    df_pno_features_sv = df_pno_features_sv[df_pno_features_sv['CustomName'].str.strip() != '']

    return df_pno_features_sv
