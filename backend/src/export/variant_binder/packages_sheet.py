from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import pandas as pd

from src.utils.db_utils import filter_df_by_timestamp, format_float_string
from src.database.db_operations import DBOperations


all_border = Border(top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'),
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'))

fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')

def get_sheet(ws, sales_versions, title, time, cell_values, config):
    """
    Fetches options data and inserts it into the specified worksheet.

    Args:
        ws (Worksheet): The worksheet to insert the data into.
        sales_versions (DataFrame): The sales versions to fetch options data for.
        title (str): The title of the sheet.

    Returns:
        None
    """
    prepare_sheet(ws, sales_versions, title, config)
    df_packages = fetch_package_data(sales_versions, time)
    
    for (code, title), group in df_packages.groupby(['Code', 'Title']):
        def get_price(x):
            if x not in group.columns:
                return ''
            if 'B' in group[x].values:
                return group[group[x] == 'B']['Price'].values[0]
            elif not group[x].dropna().empty and group[x].dropna().loc[(group[x] != 'S') & (group[x] != 'E')].empty:
                return 'S'
            else:
                return ''

        sales_versions['PKGPrice'] = sales_versions['SalesVersion'].apply(get_price)
        group = group.drop(columns=['Code', 'Title', 'Price'])
        # combine lines on column RuleCode by keeping the first non NaN value in the sales versions columns
        group = group.groupby('RuleCode').first().reset_index()
        
        df_options = group.sort_values(by='RuleCode', ascending=True)
        df_options = df_options.fillna('')

        # Write the data to the worksheet
        prices = sales_versions[(sales_versions['PKGPrice'] != '') & (sales_versions['PKGPrice'] != 'S')]['PKGPrice'].unique().tolist()

        if len(prices) == 1:
            prices = prices[0].split('/')
            ws.append([code, title, prices[0]] + sales_versions['PKGPrice'].map(lambda x: cell_values['B'] if len(x) and x[0].isnumeric() else cell_values.get(x, x)).tolist())
            ws.append(['', '', prices[1]])
        elif len(prices) == 0:
            ws.append([code, title, 'Serie'] + sales_versions['PKGPrice'].map(lambda x: cell_values['B'] if len(x) and x[0].isnumeric() else cell_values.get(x, x)).tolist())
            ws.append(['', '', ''])
        else:
            ws.append([code, title, config['PACKAGES_SHEET']['MULTI_PRICE']] + sales_versions['PKGPrice'].apply(lambda x: x.split('/')[0] if len(x) and x[0].isnumeric() else cell_values.get(x, x)).tolist())
            ws.append(['', '', ''] + sales_versions['PKGPrice'].apply(lambda x: x.split('/')[1] if len(x.split('/')) != 1 else '').tolist())
        
        for col in range(1, len(sales_versions) + 4):
            cell = ws.cell(row=len(ws["A"])-1, column=col)
            cell.fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
            cell.font = Font(name='Arial', size=12, bold=True)
            alig = 'center' if cell.column > 2 else 'left'
            cell.alignment = Alignment(horizontal=alig, vertical='center', wrap_text=True)

        # merge the cells of the 2 rows vertically besides in column C
        for col in range(1, len(sales_versions) + 4):
            if not ws.cell(row=len(ws["A"])-1, column=col).value.replace('.', '').replace(',', '').isnumeric():
                ws.merge_cells(start_row=len(ws["A"])-1, start_column=col, end_row=len(ws["A"]), end_column=col)
                ws.cell(row=len(ws["A"])-1, column=col).border = all_border
                ws.cell(row=len(ws["A"]), column=col).border = all_border
            else:
                # no bottom border
                ws.cell(row=len(ws["A"])-1, column=col).border = Border(top=Side(style='thin', color='000000'),
                                                                        left=Side(style='thin', color='000000'),
                                                                        right=Side(style='thin', color='000000'))
                ws.cell(row=len(ws["A"])-1, column=col).alignment = Alignment(horizontal='center', vertical='bottom')
                ws.cell(row=len(ws["A"]), column=col).border = Border(left=Side(style='thin', color='000000'),
                                                                        right=Side(style='thin', color='000000'),
                                                                        bottom=Side(style='thin', color='000000'))
                ws.cell(row=len(ws["A"]), column=col).alignment = Alignment(horizontal='center', vertical='top')
                ws.cell(row=len(ws["A"]), column=col).fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

        for _, row in df_options.iterrows():
            ws.append([row['RuleCode'], row['CustomName'], row['RuleBase']] + [cell_values[row[sv] if sv in df_options.columns else ''] for sv in sales_versions['SalesVersion']])
            for cell in ws[len(ws["A"])]:
                if cell.column > 2:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else: 
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    cell.border = all_border
    format_sheet(ws, len(sales_versions) + 4)

def prepare_sheet(ws, df_sales_versions, title, config):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = ws['A2']
    ws['A1'] = f'{title} - ' + config['PACKAGES_SHEET']["TITLE"]
    ws['C1'] = config['DEFAULT']['TAX'].replace('\\n', '\n')
    ws['A2'] = config['PACKAGES_SHEET']['CODE_COLUMN']
    ws['B2'] = config['PACKAGES_SHEET']['DESCRIPTION_COLUMN']
    
    # Put a thin border around the content of A2 and B2
    ws['A2'].border = all_border
    ws['B2'].border = all_border
    
    ws.merge_cells('A1:B1')
    df_sales_versions = df_sales_versions.reset_index(drop=True)
    for idx, row in df_sales_versions.iterrows():
        ws.cell(row=1, column=idx+4, value=f'{row["SalesVersionName"]}\nSV {row["SalesVersion"]}')

    # Definition of column widths & row heights
    max_c = ws.max_column

    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 65
    ws.column_dimensions['C'].width = 40
    for col in range(3, max_c + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25

    ws.row_dimensions[1].height = 45
    ws.row_dimensions[2].height = 35

    # Formatting of first row
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    
    for col in range(1,max_c+1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        if col > 2:
            cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    
    # Formatting of second row
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    ws['A2'].font = Font(size=10, bold=True)
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B2'].font = Font(size=10, bold=True)
    
def format_sheet(ws, max_c):
    # add right border to the last column
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=max_c, max_col=max_c):
        for cell in row:
            cell.border = Border(left=Side(style='thin', color='000000'))

    max_r = ws.max_row + 1
    # add bottom border to the last row
    for row in ws.iter_rows(min_row=max_r, max_row=max_r, min_col=1, max_col=max_c-1):
        for cell in row:
            cell.border = Border(top=Side(style='thin', color='000000'))

def fetch_package_data(sales_versions, time):
    pno_ids = sales_versions.ID.unique().tolist()
    conditions = []
    if len(pno_ids) == 1:
        conditions.append(f"PNOID = '{pno_ids[0]}'")
    else:
        conditions.append(f"PNOID in {tuple(pno_ids)}")
    # pno_package_conditions = conditions.copy() + ["Code = 'P0030'", "PNOID = '182CCF22-DF5A-4129-BE82-2643373A42CA'"]
    pno_package_conditions = conditions.copy() + ["Code not like 'PA%'"]
    sales_versions = sales_versions.rename(columns={'ID': 'TmpCode'})
    df_pno_package = DBOperations.instance.get_table_df(DBOperations.instance.config.get('AUTH', 'PKG'), columns=['ID', 'PNOID', 'Code', 'Title', 'RuleCode', 'RuleName', 'RuleBase', 'StartDate', 'EndDate'], conditions=pno_package_conditions)
    df_pno_package = filter_df_by_timestamp(df_pno_package, time)
    df_pno_package = df_pno_package.drop(columns=['StartDate', 'EndDate'])
    df_pno_package['RuleCode'] = df_pno_package['RuleCode'].apply(lambda x: str(x).lstrip('0') if str(x).isdigit() else x)
    
    rule_codes = df_pno_package['RuleCode'].unique().tolist()
    # rule_codes += [f"{x}  " for x in rule_codes]
    # rule_codes += [f"{x}(u)" for x in rule_codes]
    pno_features_conditions = conditions.copy()
    if len(rule_codes) == 1:
        pno_features_conditions.append(f"Reference = '{rule_codes[0]}'")
    else:
        pno_features_conditions.append(f"Reference in {tuple(rule_codes)}")
    df_pno_features_all = DBOperations.instance.get_table_df(DBOperations.instance.config.get('AUTH', 'FEAT'), columns=['PNOID', 'Code as FeatCode', 'Reference as RuleCode', 'RuleName', 'CustomName', 'StartDate'], conditions=pno_features_conditions)
    df_pno_features_all = df_pno_features_all.sort_values(by='StartDate', ascending=False)
    df_pno_features = df_pno_features_all.drop_duplicates(subset=['PNOID', 'FeatCode', 'RuleCode'], keep='first').drop(columns=['FeatCode', 'StartDate'])
    
    # Remove Options that have a subset of the relatable features
    for ref in df_pno_features['RuleCode'].unique():
        local_group = df_pno_features[df_pno_features['RuleCode'] == ref].groupby(['PNOID']).agg({'RuleName': lambda x: len(x)})
        max_num_feats = local_group['RuleName'].max()
        # Get sales versions with the maximum number of features
        max_sales_versions = local_group[local_group['RuleName'] != max_num_feats].index.tolist()
        if len(max_sales_versions):
            df_pno_features = df_pno_features[~((df_pno_features['RuleCode'] == ref) & (df_pno_features['PNOID'].isin(max_sales_versions)))]
    
    uph_codes = list(set(rule_codes) - set(df_pno_features.RuleCode.unique().tolist()))
    df_pkg_uph = pd.DataFrame()
    if len(uph_codes):
        uph_conditions = conditions.copy() + [f"StartDate <= {time}", f"EndDate >= {time}"]
        if len(uph_codes) == 1:
            uph_conditions.append(f"Code = '{uph_codes[0]}'")
        else:
            uph_conditions.append(f"Code in {tuple(uph_codes)}")
        df_uph = DBOperations.instance.get_table_df(DBOperations.instance.config.get('AUTH', 'UPH'), columns=['ID', 'PNOID', 'Code'], conditions=uph_conditions)
        df_uph = df_uph.rename(columns={'Code': 'RuleCode'})
        rel_ids = df_uph.ID.unique().tolist()
        uph_features_conditions = []
        if len(rel_ids) == 1:
            uph_features_conditions.append(f"RelationID = '{rel_ids[0]}'")
        else:
            uph_features_conditions.append(f"RelationID in {tuple(rel_ids)}")
        
        df_uph_relations = DBOperations.instance.get_table_df(DBOperations.instance.config.get('RELATIONS', 'UPH_Custom'), columns=['RelationID', 'CustomName'], conditions=uph_features_conditions)
        df_uph = df_uph.merge(df_uph_relations, left_on='ID', right_on='RelationID', how='left')
        df_uph = df_uph.drop(columns=['ID', 'RelationID'])
        df_pkg_uph = df_uph.merge(df_pno_package[['PNOID', 'RuleCode', 'RuleName']], on=['PNOID', 'RuleCode'], how='left', suffixes=('_uph', '_package'))
        df_pkg_uph = df_pkg_uph.dropna(subset=['RuleName']).drop_duplicates(subset=['PNOID', 'RuleCode', 'RuleName'])
        df_pkg_uph['CustomName'] = df_pkg_uph['CustomName'].fillna('')
    
    df_merged = df_pno_features.merge(df_pno_package[['PNOID', 'RuleCode', 'RuleName']], on=['PNOID', 'RuleCode'], how='left', suffixes=('_features', '_package'))
    df_merged['RuleName'] = df_merged['RuleName_package'].combine_first(df_merged['RuleName_features'])
    df_merged = df_merged.drop(columns=['RuleName_features', 'RuleName_package'])
    df_pno_package = df_pno_package.drop(columns=['RuleName'])
    
    df_merged = pd.concat([df_merged, df_pkg_uph], ignore_index=True)
    merged_df_with_features = df_merged.merge(df_pno_package, on=['PNOID', 'RuleCode'], how='left')
    merged_df_with_features['RuleBase'] = merged_df_with_features['RuleBase'].fillna('')
    df_filled = pd.DataFrame()
    for rule_code, rule_base, pkg_id, pkg_code, pkg_title in merged_df_with_features[['RuleCode', 'RuleBase', 'ID', 'Code', 'Title']].dropna(subset=['Code']).drop_duplicates().values:
        df_tmp = merged_df_with_features[(merged_df_with_features['RuleCode'] == rule_code) & (pd.isna(merged_df_with_features['Code']))].copy()
        if df_tmp.empty:
            continue
        df_tmp['ID'] = pkg_id
        df_tmp['Code'] = pkg_code
        df_tmp['Title'] = pkg_title
        df_tmp['RuleBase'] = rule_base
        df_filled = pd.concat([df_filled, df_tmp], ignore_index=True)
    
    merged_df_with_features = merged_df_with_features.dropna(subset=['Code'])
    merged_df_with_features = pd.concat([merged_df_with_features, df_filled], ignore_index=True)
    
    df_pno_package_with_sv = merged_df_with_features.merge(sales_versions[['TmpCode', 'SalesVersion', 'SalesVersionName']], left_on='PNOID', right_on='TmpCode', how='inner')
    df_pno_package_with_sv =df_pno_package_with_sv.drop(columns=['TmpCode', 'PNOID']).drop_duplicates()
    
    package_ids = df_pno_package_with_sv['ID'].dropna().unique().tolist()
    pno_package_price_conditions = []
    if len(package_ids) == 1:
        pno_package_price_conditions.append(f"RelationID = '{package_ids[0]}'")
    else:
        pno_package_price_conditions.append(f"RelationID in {tuple(package_ids)}")
    
    df_pno_package_price = DBOperations.instance.get_table_df(DBOperations.instance.config.get('RELATIONS', 'PKG_Custom'), columns=['RelationID', 'Price', 'PriceBeforeTax', 'CustomName as PCustomName'], conditions=pno_package_price_conditions)
    if df_pno_package_price.empty:
        df_pno_package_with_sv['RelationID'] = None
        df_pno_package_with_sv['Price'] = None
        df_pno_package_with_sv['PriceBeforeTax'] = None
        df_pno_package_with_sv['PCustomName'] = None
    else:
        df_pno_package_price = df_pno_package_price.drop_duplicates()
        df_pno_package_with_sv = df_pno_package_with_sv.merge(df_pno_package_price, left_on='ID', right_on='RelationID', how='left')
    
    df_pno_package_with_sv.loc[(df_pno_package_with_sv['Price'] == 0) & (df_pno_package_with_sv['RuleBase'] == ''), 'RuleName'] = 'S'
    new_rows = []
    ids_to_remove = []
    
    for rule_code, pkg_code in df_pno_package_with_sv[['RuleCode', 'Code']].drop_duplicates().values:
        if pd.isna(pkg_code):
            continue
        rule_text = ';\n'.join([x for x in df_pno_package_with_sv[(df_pno_package_with_sv['RuleCode'] == rule_code) & (df_pno_package_with_sv['Code'] == pkg_code)]['CustomName'].unique() if x])
        df_pno_package_with_sv.loc[(df_pno_package_with_sv['RuleCode'] == rule_code) & (df_pno_package_with_sv['Code'] == pkg_code), 'CustomName'] = rule_text
    
    df_pno_package_with_sv.loc[df_pno_package_with_sv['RuleBase']!='', 'RuleName'] = 'E'
    df = df_pno_package_with_sv
    
    for rule_code in df['RuleCode'].unique():
        template_rows = df[(df['RuleCode'] == rule_code) & df['ID'].notna() & df['Code'].notna() & df['Title'].notna() & df['RuleBase'].notna()]
        should_sales_versions = df[(df['RuleCode'] == rule_code) & (df['RuleName'] == 'S')]['SalesVersion'].unique()
        for code in template_rows['Code'].unique():
            group = df [(df['RuleCode'] == rule_code) & (df['Code'] == code)]
            available_svs = group['SalesVersion'].unique()
            not_available_sales_versions = [sv for sv in should_sales_versions if sv not in available_svs]
            for sales_version in not_available_sales_versions:
                tmp_df = df[(df['RuleCode'] == rule_code) & (df['SalesVersion'] == sales_version)]
                if tmp_df.empty:
                    continue
                if 'O' in tmp_df['RuleName'].unique():
                    ids_to_remove += tmp_df.index.tolist()
                    continue
                merged_row = tmp_df.iloc[0].to_dict()
                merged_row.update({'Code': code,
                                   'Title': group['Title'].values[0],
                                   'RuleBase': '',
                                   'PCustomName': group['PCustomName'].values[0]})
                
                new_rows.append(merged_row)
                ids_to_remove += tmp_df.index.tolist()
    
    # Convert the list of dictionaries to a DataFrame
    df_new_rows = pd.DataFrame(new_rows).drop_duplicates()
    ids_to_remove = list(set(ids_to_remove))
    df_pno_package_filtered = df_pno_package_with_sv.drop(ids_to_remove)
    
    # Append new_rows to the DataFrame
    df_pno_package_with_price = pd.concat([df_pno_package_filtered, df_new_rows], ignore_index=True)
    
    # # Step 1: Group the DataFrame
    # grouped = df_pno_package_with_price.groupby(['Code', 'SalesVersion', 'SalesVersionName']).agg({'RuleName': lambda x: set(x)})
    
    # # Step 2: Filter groups where all RuleName values are 'S'
    # filtered_groups = grouped[grouped['RuleName'].apply(lambda x: x == {'S'})].index
    
    # # Step 3: Drop these groups from the original DataFrame
    # filtered_groups = df[~df.set_index(['Code', 'SalesVersion', 'SalesVersionName']).index.isin(filtered_groups)]
    
    # Format prices to float using format_float_string
    df_pno_package_with_price['Price'] = df_pno_package_with_price['Price'].apply(format_float_string)
    df_pno_package_with_price['PriceBeforeTax'] = df_pno_package_with_price['PriceBeforeTax'].apply(format_float_string)
    
    # Concatenate Price and PriceBeforeTax
    df_pno_package_with_price['Price'] = df_pno_package_with_price.apply(lambda x: f"{x['Price']}/{x['PriceBeforeTax']}" if x['RuleName'] != '%' else '%', axis=1)
    
    # Title is custom name if available, else Title
    df_pno_package_with_price['Title'] = df_pno_package_with_price['PCustomName'].combine_first(df_pno_package_with_price['Title'])
    df_pno_package_with_price = df_pno_package_with_price.drop(columns=['ID', 'RelationID', 'PriceBeforeTax', 'PCustomName'])
    
    # Create the pivot table
    pivot_df = df_pno_package_with_price.pivot_table(index=['Code', 'Price', 'RuleCode', 'RuleBase'], columns='SalesVersion', values='RuleName', aggfunc='first')
    
    # Drop the now unneeded columns and duplicates
    df_pno_package_final = df_pno_package_with_price[['Code', 'Price', 'RuleCode', 'RuleBase', 'Title', 'CustomName']]
    df_pno_package_final = df_pno_package_final.drop_duplicates()
    
    # Join the pivoted DataFrame with the original one. sort after code ascending
    df_result = df_pno_package_final.join(pivot_df, on=['Code', 'Price', 'RuleCode', 'RuleBase']).sort_values(by='Code')
    df_result['RuleCode'] = df_result['RuleCode'].apply(lambda x: x.zfill(6))
    
    return df_result
