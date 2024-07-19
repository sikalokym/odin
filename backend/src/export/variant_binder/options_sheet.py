from openpyxl.styles import Font, PatternFill, Border, Alignment, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from src.database.db_operations import DBOperations
from src.utils.db_utils import format_float_string


#General formating of border lines & cell colours in excel spreadsheet
all_border = Border(top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'),
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'))

white_border = Border(right=Side(style='thin', color="FFFFFF"))
fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')
white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

def get_sheet(ws, sales_versions, title, time, rad_category, rule_texts, cell_values, config):
    """
    Fetches options data and inserts it into the specified worksheet.

    Args:
        ws (Worksheet): The worksheet to insert the data into.
        sales_versions (DataFrame): The sales versions to fetch options data for.
        title (str): The title of the sheet.

    Returns:
        df_rad (DataFrame): The dataframe containing the tiers data for the next sheet.
    """
    
    df_rad, df_res = fetch_options_data(sales_versions, time, rad_category, rule_texts, config)
    ws.append([''])
    ws.append([''])
    
    # insert data into the sheet ab row 3
    for _, row in df_res.iterrows():
        svs = [cell_values.get(row[sv], row[sv]) if sv in df_res.columns else '' for sv in sales_versions['SalesVersion']]
        if row['Price'] == config['OPTIONS_SHEET']['GRAYED_OUT_PRICE']or row['Price'] == config['DEFAULT']['SERIES']:
            ws.append([row['Code'], row['CustomName'], row['Price']] + svs + ['', row['CustomCategory'], row['Rules']])
            row_height = int(len(row['CustomName']) / 65) * 15 + 15
            ws.row_dimensions[ws.max_row].height = row_height
            ws.append([''])
        else:
            prices = row['Price'].split('/')
            if len(prices) > 1:
                ws.append([row['Code'], row['CustomName'], prices[0]] + svs + ['', row['CustomCategory'], row['Rules']])
                ws.cell(row=ws.max_row, column=3).alignment = Alignment(horizontal='center', vertical='bottom')
                row_height = int(len(row['CustomName']) / 65) * 15 + 15
                ws.row_dimensions[ws.max_row].height = row_height
                ws.append(['', '', prices[1]] + svs) 
                ws.cell(row=ws.max_row, column=3).alignment = Alignment(horizontal='center', vertical='top')
    
    prepare_sheet(ws, sales_versions, f'{title} - {config["OPTIONS_SHEET"]["TITLE"]}', config)

    return df_rad

def prepare_sheet(ws, df_sales_versions, title, config):
    # Finding of the last columns and rows with content for border lines
    ws.freeze_panes = ws['A2']
    max_c = ws.max_column
    max_r = ws.max_row

    for col in range(4, max_c + 1):
        for row in range(3, max_r + 1, 2):
            ws.merge_cells(start_row=row, end_row=row + 1, start_column=col, end_column=col)

    # Content of Row 1 & 2
    ws.merge_cells('A1:B1')
    ws['A1'] = title
    ws['C1'] = config['DEFAULT']['TAX'].replace('\\n', '\n')
    ws['A2'] = config['PACKAGES_SHEET']['CODE_COLUMN']
    ws['B2'] = config['PACKAGES_SHEET']['DESCRIPTION_COLUMN']
    
    # Put a thin border around the content of A2 and B2
    ws['A2'].border = all_border
    ws['B2'].border = all_border
    
    df_sales_versions = df_sales_versions.reset_index(drop=True)
    for idx, row in df_sales_versions.iterrows():
        ws.cell(row=1, column=idx+4, value=f'{row["SalesVersionName"]}\nSV {row["SalesVersion"]}')
    
    max_c = ws.max_column
    max_r = ws.max_row
    
    #Definition of column widths & row heights
    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 65
    for col in range(3, max_c + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25
    
    ws.row_dimensions[1].height = 45
    ws.row_dimensions[2].height = 43
    
    # Formatting of first row
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    
    for col in range(3, max_c+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    
    for col in range(1,max_c+1):
        cell = ws.cell(row=1, column=col)
        if col != max_c-2:
            cell.fill = fill
        else:
            cell.fill = white_fill
    
    # Formatting of second row
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['A2'].font = Font(size=10, bold=True)
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['B2'].font = Font(size=10, bold=True)

    # Formatting of first row border lines
    for col in range(1, max_c):
        cell = ws.cell(row=1, column=col)
        cell.border = white_border

    # Setting of border lines context sensitive to last cells with content
    content_area = ws.iter_cols(min_row=3, max_row=max_r, min_col=1, max_col=max_c)
    for row in content_area:
        for cell in row:
            cell.border = all_border

    for row in range(3, max_r + 1):
        for col in range(4, max_c + 1):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell = ws.cell(row=row, column=2)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell = ws.cell(row=row, column=1)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell = ws[f'C{row}']
        if row % 2 != 0:
           cell.font = Font(bold=True)
           cell.border = Border(top=Side(style='thin'),
                                right=Side(style='thin'),
                                left=Side(style='thin'))
           ws.cell(row=row-1, column=3).border = Border(bottom=Side(style='thin'),
                                                        right=Side(style='thin'),
                                                        left=Side(style='thin'))
    
    # Setting of border lines context sensitive around data
    for col in range(1, max_c + 1):
        cell = ws.cell(row=3, column=col)
        cell.border = Border(top=Side(style='medium'),
                            right=Side(style='thin'))
    
    for row in range(2, max_r + 1):
        cell = ws.cell(row=row, column=max_c)
        cell.border = Border(right=Side(style='thin'),
                            bottom=Side(style='thin'))
    
    for col in range(1, max_c + 1):
        cell = ws.cell(row=max_r, column=col)
        cell.border = Border(bottom=Side(style='thin'),
                            left=Side(style='thin'))
    
    top_right_cell = ws.cell(row=3, column=max_c)
    bottom_right_cell = ws.cell(row=max_r, column=max_c)
    
    top_right_cell.border = Border(right=Side(style='medium'),
                                        bottom=Side(style='thin'),
                                        top=Side(style='medium'),
                                        left=Side(style='thin'))
    
    bottom_right_cell.border = Border(right=Side(style='medium'),
                                        bottom=Side(style='medium'),
                                        top=Side(style='thin'),
                                        left=Side(style='thin'))
    
    # Formatting of "Pack Only" options
    for row in range(3, max_r + 1):  
        if ws.cell(row=row, column=3).value == config['OPTIONS_SHEET']['GRAYED_OUT_PRICE'] or ws.cell(row=row, column=ws.max_column-1).value == config['OPTIONS_SHEET']['GRAYED_OUT_CATEGORY']:
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')
            for col in range(1, max_c + 1):
                ws.cell(row=row, column=col).font = Font(color="A6A6A6", bold = False)
                ws.merge_cells(start_row=row, end_row=row + 1, start_column=3, end_column=3)
        elif ws.cell(row=row, column=3).value == config['DEFAULT']['SERIES']:
            ws.merge_cells(start_row=row, end_row=row+1, start_column=3, end_column=3)
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')

    row = 3
    while row <= ws.max_row:
        # Read code in current row
        current_code = ws[f"A{row}"].value

        # Search for next different code
        next_row = row + 2
        while next_row <= ws.max_row:
            next_code = ws[f"A{next_row}"].value
            if next_code != current_code:
                break
            next_row += 2

        # Merging in columns A & B if codes have duplicates
        if next_row - row > 2:
            ws.merge_cells(start_row=row, end_row=next_row-1, start_column=1, end_column=1)
            ws.merge_cells(start_row=row, end_row=next_row-1, start_column=2, end_column=2)
        else:
            ws.merge_cells(start_row=row, end_row=row+1, start_column=1, end_column=1)
            ws.merge_cells(start_row=row, end_row=row+1, start_column=2, end_column=2)

        # Set next row as first row of new control circle
        row = next_row
    
    # Formatting of gap to extra section
    ws.column_dimensions[get_column_letter(max_c - 2)].width = 1

    # iterate the column cells and set the border left and right
    for row in range(1, ws.max_row +1):
        ws.cell(row=row, column=max_c - 2).border = Border(left=Side(style='thin'),
                                                            right=Side(style='thin'))

    # Formatting of extra section
    ws.cell(row=1, column=max_c -1).value = config['OPTIONS_SHEET']['CATEGORY_COLUMN']
    ws.column_dimensions[get_column_letter(max_c-1)].width = 25

    # copy the merge range in column A to the new column
    for row in range(3, max_r + 1):
        cell = ws.cell(row=row, column=max_c-1)
        cell.border = Border(bottom=Side(style='thin'),
                            left=Side(style='thin'),
                            top=Side(style='thin'),
                            right=Side(style='thin'))
    
    # Formatting of extra section
    ws.cell(row=1, column=max_c).value = config['OPTIONS_SHEET']['NOTES']
    ws.column_dimensions[get_column_letter(max_c)].width = 25

    for row in range(3, max_r + 1):
        cell = ws.cell(row=row, column=max_c)
        cell.border = Border(bottom=Side(style='thin'),
                            left=Side(style='thin'),
                            top=Side(style='thin'),
                            right=Side(style='thin'))
    
    ws.sheet_view.showGridLines = False

def fetch_options_data(sales_versions, time, rad_category, rule_texts, config):
    pno_ids = sales_versions.ID.unique().tolist()
    conditions = [f"StartDate <= {time}", f"EndDate >= {time}"]
    if len(pno_ids) == 1:
        conditions.append(f"PNOID = '{pno_ids[0]}'")
    else:
        conditions.append(f"PNOID in {tuple(pno_ids)}")
    opt_conds = conditions.copy() + ["Code not like 'A%'"]
    df_pno_options = DBOperations.instance.get_table_df(DBOperations.instance.config.get('AUTH', 'OPT'), columns=['ID', 'PNOID', 'Code', 'RuleName', 'StartDate', 'EndDate'], conditions=opt_conds)
    if df_pno_options.empty:
        return pd.DataFrame(), pd.DataFrame()
    df_pno_options = df_pno_options.drop(columns=['StartDate', 'EndDate'])
    df_pno_options = df_pno_options.drop_duplicates()
    rel_codes = df_pno_options.ID.unique().tolist()
    pno_option_price_conditions = []
    if len(rel_codes) == 1:
        pno_option_price_conditions.append(f"RelationID = '{rel_codes[0]}'")
    else:
        pno_option_price_conditions.append(f"RelationID in {tuple(rel_codes)}")
    df_pno_option_price = DBOperations.instance.get_table_df(DBOperations.instance.config.get('RELATIONS', 'OPT_Custom'), columns=['RelationID', 'Price', 'PriceBeforeTax', 'CustomName'], conditions=pno_option_price_conditions)
    df_pno_option_price = df_pno_option_price.drop_duplicates()
    df_pno_options_with_price = df_pno_options.merge(df_pno_option_price, left_on='ID', right_on='RelationID', how='left')
    df_pno_options_with_price['OptCode'] = df_pno_options_with_price['Code'].apply(lambda x: x.lstrip('0') if x.isnumeric() else x)
    opt_codes = df_pno_options_with_price['OptCode'].unique().tolist()
    rad_title = config['TIRES_SHEET']['TITLE']
    pno_features_conditions = conditions.copy() + ["Reference != ''", f"((CustomCategory = '{rad_title}' AND CustomName != '' AND CustomName IS NOT NULL) OR (CustomCategory != '{config['TIRES_SHEET']['TITLE']}') OR (CustomCategory IS NULL))"]
    pno_rad_conditions = conditions.copy() + [f"CustomCategory = '{rad_title}'", "RuleName = 'S'", "CustomName != ''"]
    if len(opt_codes) == 1:
        pno_features_conditions.append(f"Reference = '{opt_codes[0]}'")
        pno_rad_conditions.append(f"Reference != '{opt_codes[0]}'")
    else:
        pno_features_conditions.append(f"Reference in {tuple(opt_codes)}")
        pno_rad_conditions.append(f"Reference not in {tuple(opt_codes)}")

    df_pno_features_all = DBOperations.instance.get_table_df(DBOperations.instance.config.get('AUTH', 'FEAT'), columns=['PNOID', 'Code', 'Reference', 'RuleName as FeatRule', 'CustomName as FeatName', 'CustomCategory', 'StartDate'], conditions=pno_features_conditions)
    
    df_pno_features_all = df_pno_features_all.sort_values(by='StartDate', ascending=False)
    df_pno_features = df_pno_features_all.drop_duplicates(subset=['PNOID', 'Code', 'Reference'], keep='first').drop(columns=['StartDate'])
    
    df_rad_unreferenced = DBOperations.instance.get_table_df(DBOperations.instance.config.get('AUTH', 'FEAT'), columns=['PNOID', 'Code as Reference', 'RuleName', 'CustomName', 'CustomCategory'], conditions=pno_rad_conditions)
    df_rad_with_sv = df_rad_unreferenced.merge(sales_versions, left_on='PNOID', right_on='ID', how='left').drop(columns=['PNOID', 'ID'])
    df_rad_with_sv['Price'] = config['DEFAULT']['SERIES']
    
    # Remove Options that have a subset of the relatable features
    for ref in df_pno_features['Reference'].unique():
        local_group = df_pno_features[df_pno_features['Reference'] == ref].groupby(['PNOID']).agg({'FeatRule': lambda x: len(x)})
        max_num_feats = local_group['FeatRule'].max()
        if max_num_feats == 1:
            continue
        # Get sales versions with the maximum number of features
        max_sales_versions = local_group[local_group['FeatRule'] != max_num_feats].index.tolist()
        if len(max_sales_versions):
            df_pno_features = df_pno_features[~((df_pno_features['Reference'] == ref) & (df_pno_features['PNOID'].isin(max_sales_versions)))]

    df_pno_options_with_feat_ref = df_pno_options_with_price[~df_pno_options_with_price['RelationID'].isna()].copy()
    df_pno_options_without_feat_ref = df_pno_options_with_price[df_pno_options_with_price['RelationID'].isna()].copy()
    df_pno_options_without_feat_ref.loc[:, 'Reference'] = df_pno_options_without_feat_ref['Code']
    
    pno_ids_to_drop = df_pno_options_without_feat_ref[df_pno_options_without_feat_ref['RuleName'] == '%'][['PNOID', 'Reference']].drop_duplicates()
    
    # Remove leading zeros from the Reference column
    pno_ids_to_drop['Reference'] = pno_ids_to_drop['Reference'].apply(lambda x: x.lstrip('0'))
    df_pno_options_feat_merged = df_pno_features.merge(df_pno_options_with_feat_ref, 
                                      how='left',
                                      left_on=['PNOID', 'Reference'],
                                      right_on=['PNOID', 'OptCode'])
    
    # Drop the rows with the same PNOID and Reference as the ones in pno_ids_to_drop
    df_pno_options_feat_merged = df_pno_options_feat_merged[~df_pno_options_feat_merged[['PNOID', 'Reference']].apply(tuple, axis=1).isin(pno_ids_to_drop.apply(tuple, axis=1))]
    
    df_pno_options_merged = pd.concat([df_pno_options_feat_merged, df_pno_options_without_feat_ref], ignore_index=True)
    
    sales_versions = sales_versions.rename(columns={'ID': 'TmpCode'})
    df_pno_options_merged = df_pno_options_merged.merge(sales_versions[['TmpCode', 'SalesVersion', 'SalesVersionName']], left_on='PNOID', right_on='TmpCode', how='left')
    
    # Replace NaN or empty string in 'CustomName' with values from 'FeatName'
    missing_or_empty_custom_name = df_pno_options_merged['CustomName'].isna() | (df_pno_options_merged['CustomName'] == '')
    df_pno_options_merged.loc[missing_or_empty_custom_name, 'CustomName'] = df_pno_options_merged.loc[missing_or_empty_custom_name, 'FeatName']

    # Replace NaN or empty string in 'RuleName' with values from 'FeatRule'
    missing_or_empty_rule_name = df_pno_options_merged['RuleName'].isna() | (df_pno_options_merged['RuleName'] == '')
    df_pno_options_merged.loc[missing_or_empty_rule_name, 'RuleName'] = df_pno_options_merged.loc[missing_or_empty_rule_name, 'FeatRule']
    
    df_pno_options_merged = df_pno_options_merged.drop(columns=['PNOID', 'OptCode', 'Code', 'RelationID', 'ID', 'TmpCode', 'FeatRule', 'FeatName']).drop_duplicates()
    
    # Append zeros to the reference column
    df_pno_options_merged['Reference'] = df_pno_options_merged['Reference'].apply(lambda x: x.zfill(6) if x.isnumeric() else x)

    def get_prices(row):
        if row['RuleName'] == 'P':
            # Get all rows with the same code that are not 'P'
            rows = df_pno_options_merged[(df_pno_options_merged['Reference'] == row['Reference']) & (df_pno_options_merged['RuleName'] != 'P')]
            # If there are no other rows with the same code, return 'Pack Only'
            if rows.empty:
                return [config['OPTIONS_SHEET']['GRAYED_OUT_PRICE']]
            # return a list of all prices
            if not rows['RuleName'].str.contains('O').any():
                return [config['OPTIONS_SHEET']['GRAYED_OUT_PRICE']]
        
        elif row['RuleName'] == '%':
            # Get all rows with the same code that are not '%'
            rows = df_pno_options_merged[(df_pno_options_merged['Reference'] == row['Reference']) & (df_pno_options_merged['RuleName'] != '%')]
            # If there are no other rows with the same code, return 'Serie'
            if rows.empty:
                return [config['DEFAULT']['SERIES']]

            return ['/'.join([f"{r['Price']}/{r['PriceBeforeTax']}" for _, r in rows.iterrows()])]
        
        elif row['RuleName'] == 'S':
            # Get all rows with the same code that are not 'S'
            rows = df_pno_options_merged[(df_pno_options_merged['Reference'] == row['Reference']) & (df_pno_options_merged['RuleName'] != 'S')]
            # If there are no other rows with the same code, return 'Serie'
            if rows.empty:
                # Should not happen since this would be a feature and not an option
                return [config['DEFAULT']['SERIES']]
            
            if not rows['RuleName'].str.contains('O').any():
                return [config['OPTIONS_SHEET']['GRAYED_OUT_PRICE']]
            
            rows = rows.drop_duplicates(subset=['Price', 'PriceBeforeTax'])
            
            # get a list of all prices in those rows
            prices = [f"{r['Price']}/{r['PriceBeforeTax']}" for _, r in rows.iterrows()]
            return prices

        return [f"{row['Price']}/{row['PriceBeforeTax']}"]

    # format prices to float using format_float_string
    df_pno_options_merged['Price'] = df_pno_options_merged['Price'].apply(format_float_string)
    df_pno_options_merged['PriceBeforeTax'] = df_pno_options_merged['PriceBeforeTax'].apply(format_float_string)
    
    # Concatenate Price and PriceBeforeTax
    df_pno_options_merged['Price'] = df_pno_options_merged.apply(lambda x: get_prices(x), axis=1)
    df_pno_options_merged = df_pno_options_merged.drop(['PriceBeforeTax'], axis=1)
    df_pno_options_merged = df_pno_options_merged.explode('Price')
    
    # Group by 'Reference' and 'Price', aggregate 'CustomName' and 'CustomCategory' columns
    aggregated = df_pno_options_merged.groupby(['Reference', 'Price']).agg({
        'CustomName': lambda x: ';\n'.join(sorted([y for y in x.dropna().unique() if y])),
        'CustomCategory': lambda x: ';\n'.join(sorted([y for y in x.dropna().unique() if y]))
    }).reset_index()
    df_pno_options_merged = df_pno_options_merged.drop(columns=['CustomName', 'CustomCategory'])
    
    # Merge the aggregated columns back to the original dataframe
    df_merged = df_pno_options_merged.merge(aggregated, on=['Reference', 'Price'])
    
    # group by Reference, Price, RuleName, SalesVersion and SalesVersionName and aggregate the CustomName column by concatinating the values with a newline separator if they differ and not null nan or empty and the CustomCategory column by concatinating the values with a comma separator if they differ
    df_pno_options_merged = df_merged.groupby(['Reference', 'Price', 'RuleName', 'SalesVersion', 'SalesVersionName']).agg({'CustomName': 'first', 'CustomCategory': 'first'}).reset_index()
    
    df_pno_options_merged = pd.concat([df_pno_options_merged, df_rad_with_sv], ignore_index=True)
    
    # Create the pivot table
    pivot_df = df_pno_options_merged.pivot_table(index=['Reference', 'Price'], columns='SalesVersion', values='RuleName', aggfunc='first')
    
    # Drop the now unneeded columns and duplicates
    df_pno_options_merged = df_pno_options_merged.drop(['RuleName', 'SalesVersion', 'SalesVersionName'], axis=1)
    df_pno_options_merged = df_pno_options_merged.drop_duplicates(['Reference', 'Price', 'CustomCategory'], keep='first')
    
    # Join the pivoted DataFrame with the original one. sort after code ascending
    df_result = df_pno_options_merged.join(pivot_df, on=['Reference', 'Price']).sort_values(by='Reference')
    
    # rename Reference to Code
    df_result = df_result.rename(columns={'Reference': 'Code'})
    
    # fill df_res columns for sv in sales_versions['SalesVersion'] with '-' if nan
    for sv in sales_versions['SalesVersion']:
        df_result[sv] = df_result[sv].fillna('-')
    
    # get the df with the rules for the options
    opt_codes = df_result['Code'].unique().tolist()
    rules_conditions = conditions.copy() + ["RuleCode LIKE 'O%'"]
    if len(opt_codes) == 1:
        rules_conditions.append(f"ItemCode = '{opt_codes[0]}'")
    else:
        rules_conditions.append(f"ItemCode in {tuple(opt_codes)}")
    df_rules = DBOperations.instance.get_table_df(DBOperations.instance.config.get('DEPENDENCIES', 'OFO'), columns=['RuleCode', 'ItemCode', 'FeatureCode'], conditions=rules_conditions)
    df_rules = df_rules.drop_duplicates()
    
    df_result['Rules'] = ''
    if df_rules.empty:
        df_rader = df_result[df_result['CustomCategory'] == config['TIRES_SHEET']['TITLE']]
        df_result = df_result[df_result['CustomCategory'] != config['TIRES_SHEET']['TITLE']]
        
        return df_rader, df_result
    
    # strip the whitespace from the ItemCode and FeatureCode columns
    df_rules['ItemCode'] = df_rules['ItemCode'].str.strip()
    df_rules['FeatureCode'] = df_rules['FeatureCode'].str.strip()
    
    for rule, group in df_rules.groupby('RuleCode'):
        group = group.groupby('ItemCode').agg({'FeatureCode': lambda x: rule_texts.get(rule.lower(), rule) + ' ' + ', '.join(list(x))}).reset_index()
        group = group.rename(columns={'FeatureCode': rule})
        df_result = pd.merge(df_result, group, left_on='Code', right_on='ItemCode', how='left')
        df_result = df_result.drop(columns=['ItemCode'])
        # accumulate the rules in the Rules column with a new line separator
        df_result['Rules'] = df_result.apply(lambda row: row['Rules'] + '\n' + row[rule] if pd.notnull(row[rule]) else row['Rules'], axis=1)
        df_result = df_result.drop(columns=[rule])
    
    # remove the first new line separator
    df_result['Rules'] = df_result['Rules'].apply(lambda x: x[1:] if x.startswith('\n') else x)
    # split the df after the CustomCategory column where it is equal Räder and return the two dataframes
    df_rader = df_result[df_result['CustomCategory'] == rad_category]
    df_result = df_result[df_result['CustomCategory'] != rad_category]
    
    return df_rader, df_result
