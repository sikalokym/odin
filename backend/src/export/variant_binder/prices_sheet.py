from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from src.database.db_operations import DBOperations
from src.utils.db_utils import filter_df_by_timestamp, format_float_string


black_side = Side(color='000000', border_style='thin')
all_border = Border(left=black_side, right=black_side, bottom=black_side, top=black_side)
blr_border = Border(left=black_side, right=black_side, bottom=black_side)
tlr_border = Border(left=black_side, right=black_side, top=black_side)


def get_sheet(ws, df_valid_pnos, df_sales_versions, title, time, df_engines_types, country):
    """
    Generates a sheet with price data for different engine types.

    Args:
        ws (worksheet): The worksheet object to write the data to.
        df_valid_pnos (DataFrame): The DataFrame containing valid part numbers.
        df_sales_versions (DataFrame): The DataFrame containing sales versions data.
        title (str): The title of the sheet.
        time (str): The time of data retrieval.
        df_engines_types (DataFrame): The DataFrame containing engine types data.

    Returns:
        None
    """
    df_price, df_gb, df_en, gb_ids = fetch_vb_price_data(country, df_valid_pnos, time)
    
    #get length of the code column, where code is an array and we want the max length of any array in the Code column
    max_code_length = df_engines_types['Code'].apply(lambda x: len(x)).max()
    prepare_sheet(ws, title, max_code_length)
    curr_row = 3
    for _, row in df_engines_types.iterrows():
        type = row['EngineType']
        group = df_price[df_price['Engine'].isin(row['Code'])]
        
        df_group_gb = pd.DataFrame()
        df_group_gb['Code'] = group['Gearbox'].drop_duplicates()
        df_group_gb['MarketText'] = df_group_gb['Code'].apply(lambda x: df_gb[df_gb['Code'] == x].iloc[0]['CustomName'])
        
        df_group_en = pd.DataFrame()
        df_group_en['Code'] = group['Engine'].drop_duplicates()
        df_group_en['MarketText'] = df_group_en['Code'].apply(lambda x: df_en[df_en['Code'] == x].iloc[0]['CustomName'])
        df_group_en['Performance'] = df_group_en['Code'].apply(lambda x: df_en[df_en['Code'] == x].iloc[0]['Performance'])
        curr_row = insert_engines_type_title(ws, type, curr_row)
        curr_row = insert_table(ws, group, df_sales_versions, df_group_gb, df_group_en, curr_row)
        curr_row += 1
    
    return gb_ids

def prepare_sheet(ws, title, max_code_length):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.headingsVisible = False
    ws.freeze_panes = ws['A2']
    for i in range(1, 100):
        ws.column_dimensions[get_column_letter(i)].width = 20
        ws.row_dimensions[i].height = 12
    ws.column_dimensions['A'].width = 45
    
    create_title(ws, f'Volvo {title} - Preise', max_code_length)

def create_title(ws, title, max_code_length):
    ws.row_dimensions[1].height = 40
    letter = get_column_letter(max_code_length + 1)
    ws.merge_cells(f'A1:{letter}1')
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=1, column=1).fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')
    ws.cell(row=1, column=1).font = Font(name='Arial', sz=18, bold=True, color='FFFFFF')

def insert_engines_type_title(ws, type, curr_row):
    ws.cell(row=curr_row, column=1, value=type)
    ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=26)
    ws.row_dimensions[curr_row].height = 24
    ws.cell(row=curr_row, column=1).font = Font(name='Arial', sz=18, bold=True)
    return curr_row + 1

def insert_table(ws, df_price, df_sv, df_gb, df_en, curr_row):
    mwst_row = curr_row
    df = df_price.copy()
    prepend_string = df.Model.iloc[0] + ' SV '
    df_sv['FullCode'] = prepend_string + df_sv['SalesVersion']
    insert_meta_column(ws, curr_row, df_sv)
    curr_row += 1
    ws.row_dimensions[curr_row].height = 24

    curr_col = 2
    for en in df.Engine.unique():
        for gb in df[df['Engine'] == en].Gearbox.unique():
            group = df[(df['Engine'] == en) & (df['Gearbox'] == gb)]
            gb_name = df_gb[df_gb['Code'] == gb].iloc[0]['MarketText']
            en_name = df_en[df_en['Code'] == en].iloc[0]['MarketText']
            en_performance = df_en[df_en['Code'] == en].iloc[0]['Performance']
            values = get_price_column(group, df_sv)
            col = [en_name, gb_name, en_performance] + values + [en, gb]
            curr_col, curr_row_end = insert_column(ws, col, curr_row, curr_col)
    insert_mwst_line(ws, mwst_row, curr_col -1)
    for i in [mwst_row+1, mwst_row+2, mwst_row+3]:
        for j in range(1, curr_col):
            ws.cell(row=i, column=j).fill = PatternFill(start_color='bfbfbf', end_color='bfbfbf', fill_type='solid')
    return curr_row_end

def insert_mwst_line(ws, curr_row, curr_col):
    rich = CellRichText(
        TextBlock(InlineFont(b=True), 'EUR inkl. 19% MwSt.'),
        '\nEUR ohne MwSt.'
    )
    # Assign the RichText to the cell
    
    ws.row_dimensions[curr_row].height = 28
    ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=curr_col)
    ws.cell(row=curr_row, column=1).value = rich
    ws.cell(row=curr_row, column=1).font = Font(name='Arial',bold=True, sz=10)
    ws.cell(row=curr_row, column=1).alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    for i in range(1, curr_col+1):
        ws.cell(row=curr_row, column=i).border = all_border
    ws.cell(row=curr_row, column=1).fill = PatternFill(start_color='bfbfbf', end_color='bfbfbf', fill_type='solid')

def get_price_column(group, df_sv):
    prices = [] 
    for sv in df_sv['SalesVersion']:
        entry = group[group['SalesVersion'] == sv]
        if entry.empty:
            prices.append('-')
            prices.append('-')
        else:
            price = entry['Price'].iloc[0]
            price_before_tax = entry['PriceBeforeTax'].iloc[0]
            prices.append(format_float_string(price))
            prices.append(format_float_string(price_before_tax))
    return prices

def insert_meta_column(ws, curr_row, df_sales_version):

    meta_fields = ['Motor', 'Getriebe', 'Leistung kW (PS)']
    ws.row_dimensions[curr_row].height = 24
    ws.row_dimensions[curr_row+1].height = 28
    ws.row_dimensions[curr_row+2].height = 12
    for i, field in enumerate(meta_fields):
        curr_row += 1
        ws.cell(row=curr_row, column=1, value=field)
        ws.cell(row=curr_row, column=1).font = Font(name='Arial',bold=True, sz=10)
        ws.cell(row=curr_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=curr_row, column=1).border = blr_border
        
    alternating_list = [item for pair in zip(df_sales_version['SalesVersionName'], df_sales_version['FullCode']) for item in pair]
    for i, field in enumerate(alternating_list):
        curr_row += 1
        ws.cell(row=curr_row, column=1, value=field)
        ws.cell(row=curr_row, column=1).font = Font(name='Arial',bold=True, sz=8)
        if i % 2:
            ws.cell(row=curr_row, column=1).border = blr_border
        else:
            ws.cell(row=curr_row, column=1).border = Border(left=black_side, right=black_side)

    footer_meta_tables = ['Motor-Codes', 'Getriebe-Codes']

    for i, field in enumerate(footer_meta_tables):
        curr_row += 1
        ws.cell(row=curr_row, column=1, value=field)
        ws.cell(row=curr_row, column=1).font = Font(name='Arial',bold=True, sz=10)
        ws.cell(row=curr_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=curr_row, column=1).fill = PatternFill(start_color='bfbfbf', end_color='bfbfbf', fill_type='solid')
        ws.cell(row=curr_row, column=1).border = blr_border

def insert_column(ws, col, curr_row, curr_col):
    for i, field in enumerate(col):
        ws.cell(row=curr_row, column=curr_col, value=field)
        ws.cell(row=curr_row, column=curr_col).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=curr_row, column=curr_col).font = Font(name='Arial', sz=8)
        ws.cell(row=curr_row, column=curr_col).border = all_border
        if i == 0:
            ws.cell(row=curr_row, column=curr_col).font = Font(name='Arial', sz=10, bold=True)
        elif i > 2:
            if i%2:
                ws.cell(row=curr_row, column=curr_col).font = Font(name='Arial',bold=True, sz=8)
                ws.cell(row=curr_row, column=curr_col).border = tlr_border
            else:
                ws.cell(row=curr_row, column=curr_col).border = blr_border
        curr_row += 1
    ws.cell(row=curr_row-2, column=curr_col).border = all_border
    return curr_col + 1, curr_row

def fetch_vb_price_data(country, df_valid_pnos, time):
    
    df_vb_prices = DBOperations.instance.get_table_df(DBOperations.instance.config.get('RELATIONS', 'PNO_Custom'), columns=['RelationID', 'Price', 'PriceBeforeTax'])
    df_price = df_valid_pnos.merge(df_vb_prices, left_on='ID', right_on='RelationID', how='left')
    
    df_gb = DBOperations.instance.get_table_df(DBOperations.instance.config.get('TABLES', 'G'), conditions=[f'CountryCode = {country}'])
    df_gb = df_gb[df_gb['Code'].isin(df_price['Gearbox'].tolist())]
    df_gb = filter_df_by_timestamp(df_gb, time)
    df_gb['CustomName'] = df_gb['CustomName'].combine_first(df_gb['MarketText'])
    gb_ids = df_gb['ID'].unique().tolist() 
    df_gb = df_gb[['Code', 'CustomName']]

    en_codes = df_price['Engine'].unique().tolist()
    conditions = [f'CountryCode = {country}']
    if len(en_codes) == 1:
        conditions.append(f'Code = {en_codes[0]}')
    else:
        conditions.append(f'Code in {tuple(en_codes)}')
    df_en = DBOperations.instance.get_table_df(DBOperations.instance.config.get('TABLES', 'En'), conditions=conditions)
    df_en = filter_df_by_timestamp(df_en, time)
    
    # sort Code by en_codes
    df_en['Code'] = pd.Categorical(df_en['Code'], en_codes)
    df_en.sort_values(by='Code', inplace=True)

    df_en['CustomName'] = df_en['CustomName'].combine_first(df_en['MarketText'])
    df_en = df_en[['Code', 'CustomName', 'Performance']]

    return df_price, df_gb, df_en, gb_ids
