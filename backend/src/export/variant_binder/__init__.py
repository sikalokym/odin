from openpyxl import Workbook
from io import BytesIO
import configparser
import pandas as pd
import numpy as np

from src.export.variant_binder import prices_sheet, options_sheet, upholstery_colors_sheet, packages_sheet, sales_versions_sheet, tiers_sheet, change_log
from src.utils.db_utils import filter_df_by_timestamp, filter_model_year_by_translation, get_model_year_from_date
from src.database.db_operations import DBOperations

# @author Hassan Wahba

def extract_variant_binder_pnos(country, model, engines_types, time):
    try:
        valid_engines = get_valid_engines(country, engines_types, time)
        valid_pnos = get_valid_pnos(country, model, time, valid_engines)
        sales_versions = get_sales_versions(country, valid_pnos, time)
        sales_versions = sales_versions[['SalesVersion', 'SalesVersionName']].drop_duplicates()
        sv_order = sales_versions['SalesVersion'].tolist()
        
        codes = valid_pnos['Model'].unique().tolist()
        conditions = [f"CountryCode = '{country}'", f'StartDate <= {time}', f'EndDate >= {time}']
        if len(codes) == 1:
            conditions.append(f"Code = '{codes[0]}'")
        else:
            conditions.append(f"Code in {tuple(codes)}")
        
        df_models = DBOperations.instance.get_table_df(DBOperations.instance.config.get('TABLES', 'Typ'), columns=['Code', 'CustomName', 'StartDate', 'EndDate'], conditions=conditions)
        df_models = filter_model_year_by_translation(df_models, conditional_columns=['CustomName'])
        df_models = df_models.drop(columns=['StartDate', 'EndDate'], axis=1)
        df_models = df_models.rename(columns={'Code': 'PNOCode'})
        
        df_pnos = valid_pnos.merge(df_models, how='left', left_on='Model', right_on='PNOCode')
        df_pnos = df_pnos.drop(columns=['PNOCode']).drop_duplicates()
        
        # Merge the sales versions with the pnos
        df_pnos = df_pnos.merge(sales_versions, how='left', on='SalesVersion')
        
        # Sort the pnos by the sales version code order
        df_pnos['SalesVersion'] = pd.Categorical(df_pnos['SalesVersion'], categories=sv_order, ordered=True)
        df_pnos = df_pnos.sort_values('SalesVersion')
        
        return df_pnos
    except Exception as e:
        DBOperations.instance.logger.error(f"Error getting VB Data: {e}", extra={'country': country})
        raise Exception(f"Error getting VB Data: {e}")
    
def extract_variant_binder(country, model, engines_types, time, pno_ids=None, sv_order=None):
    config = configparser.ConfigParser()
    config.read(f'config/variant_binder/{country}.cfg', encoding='utf-8')
    cell_values = dict()
    for x in config['CELL_VALUES']:
        symbols = config['CELL_VALUES'][x].split(',')
        for symbol in symbols:
            cell_values[symbol] = x
    uph_col_cell_values = dict()
    for x in config['CELL_VALUES_UPH_COL']:
        symbols = config['CELL_VALUES_UPH_COL'][x].split(',')
        for symbol in symbols:
            uph_col_cell_values[symbol] = x
    rule_texts = dict(config['RULES_TEXTS'])
    wb = Workbook()
    
    # Remove the default sheet created
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    try:
        valid_engines = get_valid_engines(country, engines_types, time)
        valid_pnos = get_valid_pnos(country, model, time, valid_engines)
    #     if pno_ids:
    #         valid_pnos = valid_pnos[valid_pnos['ID'].isin(pno_ids)]
        
    #     sales_versions = get_sales_versions(country, valid_pnos, time)
    #     if sv_order:
    #         sales_versions['SalesVersion'] = pd.Categorical(sales_versions['SalesVersion'], categories=sv_order, ordered=True)
    #         sales_versions = sales_versions.sort_values('SalesVersion')
    #         sales_versions['SalesVersion'] = sales_versions['SalesVersion'].astype(str)
    #     sales_versions = sales_versions[['ID', 'SalesVersion', 'SalesVersionName']].drop_duplicates(subset='SalesVersion')
    #     title, model_id = get_model_name(country, model, time)
    except Exception as e:
        DBOperations.instance.logger.error(f"Error getting VB Data: {e}", extra={'country': country})
        raise Exception(f"Error getting VB Data: {e}")
    
    # if valid_pnos.empty or sales_versions.empty or valid_engines.empty:
    #     DBOperations.instance.logger.warning(f"No data found for model {model} and engine category {engines_types} at time {time}", extra={'country': country})
    #     return None
    # try:
    #     ws_1 = wb.create_sheet(config['PRICES_SHEET']['TITLE'])
    #     gb_ids = prices_sheet.get_sheet(ws_1, valid_pnos, sales_versions.copy(), title, time, valid_engines, country, config['DEFAULT']['TAX'].replace('\\n', '\n'), dict(config['PRICES_SHEET']))
    # except Exception as e:
    #     DBOperations.instance.logger.error(f"Error creating sheet prices: {e}", extra={'country': country})
    # try:
    #     ws_2 = wb.create_sheet(config['FEATURES_SHEET']['TITLE'])
    #     sales_versions_sheet.get_sheet(ws_2, sales_versions.copy(), title, dict(config['FEATURES_SHEET']))
    # except Exception as e:
    #     DBOperations.instance.logger.error(f"Error creating sheet features: {e}", extra={'country': country})
    # try:
    #     ws_3 = wb.create_sheet(config['PACKAGES_SHEET']['TITLE'])
    #     packages_sheet.get_sheet(ws_3, sales_versions.copy(), title, time, cell_values, config)
    # except Exception as e:
    #     DBOperations.instance.logger.error(f"Error creating sheet: {e}", extra={'country': country})
    # try:
    #     ws_4 = wb.create_sheet(config['UPHOLSTERY_COLORS_SHEET']['TITLE'])
    #     upholstery_colors_sheet.get_sheet(ws_4, sales_versions.copy(), title, time, uph_col_cell_values, rule_texts, config)
    # except Exception as e:
    #     DBOperations.instance.logger.error(f"Error creating sheet upholstery colors: {e}", extra={'country': country})
    # df_rad = None
    # try:
    #     ws_5 = wb.create_sheet(config['OPTIONS_SHEET']['TITLE'])
    #     df_rad = options_sheet.get_sheet(ws_5, sales_versions.copy(), title, time, config['TIRES_SHEET']['TITLE'], rule_texts, cell_values, config)
    # except Exception as e:
    #     DBOperations.instance.logger.error(f"Error creating sheet options: {e}", extra={'country': country})
    # if df_rad is not None:
    #     try:
    #         ws_6 = wb.create_sheet(config['TIRES_SHEET']['TITLE'])
    #         tiers_sheet.get_sheet(ws_6, sales_versions.copy(), title, df_rad, cell_values, config)
    #     except Exception as e:
    #         DBOperations.instance.logger.error(f"Error creating sheet tires: {e}", extra={'country': country})
    
    try:
        ws_7 = wb.create_sheet(config['CHANGES_SHEET']['TITLE'], 0)
        # entities_ids_dict = {'Typ': [model_id], 'SV': sales_versions.ID.unique().tolist(), 'En': valid_engines.ID.explode().unique().tolist(), 'G': gb_ids}
        err = change_log.get_sheet(ws_7, None, None, None, time, country)
        if err:
            raise Exception('No data found for change log')
    except Exception as e:
        DBOperations.instance.logger.error(f"Error creating sheet: {e}", extra={'country': country})

    wb.active = wb.index(ws_1)
    model_year = get_model_year_from_date(time)
    time = str(time)
    vb_title = f"{title.replace(' ', '')}_VB_{engines_types}_{model_year}_{time[:4]}w{time[4:]}.xlsx"
    # wb.save(f"dist/vbs/{vb_title}")
    # wb.save(f"{vb_title}") # for testing
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output, vb_title

def _extract_variant_binder(country, model, engines_types, time, pno_ids=None, sv_order=None):
    config = configparser.ConfigParser()
    config.read(f'config/variant_binder/{country}.cfg', encoding='utf-8')
    # config.read(f'config/variant_binder/template.cfg', encoding='utf-8')
    cell_values = dict()
    for x in config['CELL_VALUES']:
        symbols = config['CELL_VALUES'][x].split(',')
        for symbol in symbols:
            cell_values[symbol] = x
    uph_col_cell_values = dict()
    for x in config['CELL_VALUES_UPH_COL']:
        symbols = config['CELL_VALUES_UPH_COL'][x].split(',')
        for symbol in symbols:
            uph_col_cell_values[symbol] = x
    rule_texts = dict(config['RULES_TEXTS'])
    wb = Workbook()
    
    # Remove the default sheet created
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    valid_engines = get_valid_engines(country, engines_types, time)
    valid_pnos = get_valid_pnos(country, model, time, valid_engines)
    if pno_ids:
        valid_pnos = valid_pnos[valid_pnos['ID'].isin(pno_ids)]
    
    sales_versions = get_sales_versions(country, valid_pnos, time)
    if sv_order:
        sales_versions['SalesVersion'] = pd.Categorical(sales_versions['SalesVersion'], categories=sv_order, ordered=True)
        sales_versions = sales_versions.sort_values('SalesVersion')
        sales_versions['SalesVersion'] = sales_versions['SalesVersion'].astype(str)
    sales_versions = sales_versions[['ID', 'SalesVersion', 'SalesVersionName']].drop_duplicates(subset='SalesVersion')
    title, model_id = get_model_name(country, model, time)
 
    if valid_pnos.empty or sales_versions.empty or valid_engines.empty:
        DBOperations.instance.logger.warning(f"No data found for model {model} and engine category {engines_types} at time {time}", extra={'country': country})
        return None
    ws_1 = wb.create_sheet(config['PRICES_SHEET']['TITLE'])
    gb_ids = prices_sheet.get_sheet(ws_1, valid_pnos, sales_versions.copy(), title, time, valid_engines, country, config['DEFAULT']['TAX'].replace('\\n', '\n'), dict(config['PRICES_SHEET']))
    ws_2 = wb.create_sheet(config['FEATURES_SHEET']['TITLE'])
    sales_versions_sheet.get_sheet(ws_2, sales_versions.copy(), title, dict(config['FEATURES_SHEET']))
    ws_3 = wb.create_sheet(config['PACKAGES_SHEET']['TITLE'])
    packages_sheet.get_sheet(ws_3, sales_versions.copy(), title, time, cell_values, config)
    ws_4 = wb.create_sheet(config['UPHOLSTERY_COLORS_SHEET']['TITLE'])
    upholstery_colors_sheet.get_sheet(ws_4, sales_versions.copy(), title, time, uph_col_cell_values, rule_texts, config)
    ws_5 = wb.create_sheet(config['OPTIONS_SHEET']['TITLE'])
    df_rad = options_sheet.get_sheet(ws_5, sales_versions.copy(), title, time, config['TIRES_SHEET']['TITLE'], rule_texts, cell_values, config)
    ws_6 = wb.create_sheet(config['TIRES_SHEET']['TITLE'])
    tiers_sheet.get_sheet(ws_6, sales_versions.copy(), title, df_rad, cell_values, config)
    ws_7 = wb.create_sheet(config['CHANGES_SHEET']['TITLE'], 0)
    entities_ids_dict = {'Typ': [model_id], 'SV': sales_versions.ID.unique().tolist(), 'En': valid_engines.ID.explode().unique().tolist(), 'G': gb_ids}
    err = change_log.get_sheet(ws_7, entities_ids_dict, valid_pnos.ID.unique().tolist(), title, time, country)
    if err:
        raise Exception('No data found for change log')
    wb.active = wb.index(ws_1)
    model_year = get_model_year_from_date(time)
    time = str(time)
    vb_title = f"{title.replace(' ', '')}_VB_{engines_types}_{model_year}_{time[:4]}w{time[4:]}.xlsx"
    wb.save(f"dist/vbs/{vb_title}")

def get_model_name(country, model, time):
    models = DBOperations.instance.get_table_df(DBOperations.instance.config.get('TABLES', 'Typ'), ['ID', 'MarketText', 'CustomName', 'StartDate', 'EndDate'], conditions=[f"CountryCode = '{country}'", f"Code = '{model}'"])

    # Filter models where StartDate and End Data wrap the current time for the given model
    df_model = filter_df_by_timestamp(models, time)

    first_row = df_model.iloc[0].copy()
    first_row['Title'] = first_row['CustomName'] if first_row['CustomName'] else first_row['MarketText']
    if not first_row.empty:
        return first_row['Title'], first_row['ID']
    else:
        raise Exception(f"No model found for model {model}")

def get_sales_versions(country, pnos, time):
    # Extract unique sales version codes
    sv_codes = pnos['SalesVersion'].unique()
    
    # Construct conditions for querying sales versions
    conditions = [
        f"CountryCode = '{country}'",
        f"StartDate <= '{time}'",
        f"EndDate >= '{time}'",
        f"Code in {tuple(sv_codes)}" if len(sv_codes) > 1 else f"Code = '{sv_codes[0]}'"
    ]
    
    # Retrieve sales version data
    df_sv = DBOperations.instance.get_table_df(
        DBOperations.instance.config.get('TABLES', 'SV'), conditions=conditions)
    df_sv = df_sv.rename(columns={'Code': 'TmpCode', 'ID': 'SVID'})

    # Merge with PNO data
    df_allowed_sv = pnos.merge(df_sv[['SVID', 'TmpCode', 'MarketText', 'CustomName']],
                               left_on='SalesVersion', right_on='TmpCode', how='left')
    df_allowed_sv['SalesVersionName'] = df_allowed_sv['CustomName'].combine_first(df_allowed_sv['MarketText'])

    # Define a function for custom price aggregation
    def custom_price_aggregation(prices):
        if len(prices) == 1:
            return prices.iloc[0]
        return ', '.join(map(str, sorted(set(prices)))) + " Check Visa File"

    # Retrieve and handle price data
    pno_condition = [f"StartDate <= '{time}'", f"EndDate >= '{time}'"]
    df_pno_price = DBOperations.instance.get_table_df(
        DBOperations.instance.config.get('RELATIONS', 'PNO_Custom'), conditions=pno_condition)
    
    if df_pno_price.empty:
        df_allowed_sv['SalesVersionPrice'] = None
    else:
        aggregated_prices = df_pno_price.groupby('RelationID')['Price'].agg(custom_price_aggregation)
        df_allowed_sv = df_allowed_sv.set_index('ID')
        df_allowed_sv['SalesVersionPrice'] = df_allowed_sv.index.map(aggregated_prices)
        df_allowed_sv.reset_index(inplace=True)

    # Clean up and reformat data
    df_allowed_sv = df_allowed_sv[['ID', 'SalesVersion', 'SalesVersionName', 'SalesVersionPrice', 'SVID']]
    df_allowed_sv['SalesVersionPrice'].fillna(value=np.nan, inplace=True)
    df_allowed_sv['SalesVersionName'].replace('', np.nan, inplace=True)
    df_allowed_sv['SalesVersionName'].fillna(df_allowed_sv['SalesVersion'], inplace=True)

    # Group and sort by the first word in SalesVersionName and max price
    df_allowed_sv['SalesVersionNameGroup'] = df_allowed_sv['SalesVersionName'].str.split().str[0]
    df_grouped = df_allowed_sv.groupby('SalesVersionNameGroup').agg({
        'SalesVersionName': list, 'SalesVersionPrice': 'max'
    }).sort_values('SalesVersionPrice', ascending=True)

    sorted_sv_names = df_grouped.explode('SalesVersionName')['SalesVersionName'].unique()
    df_final = df_allowed_sv.set_index('SalesVersionName').loc[sorted_sv_names].reset_index()
    df_final = df_final.drop(columns=['SalesVersionNameGroup'])
    df_final['SalesVersionName'] = df_final.apply(
        lambda x: '' if x['SalesVersionName'] == x['SalesVersion'] else x['SalesVersionName'], axis=1)

    return df_final.reset_index(drop=True)

def get_valid_engines(country, engine_cat, time):
    df_all_engines = DBOperations.instance.get_table_df(DBOperations.instance.config.get('TABLES', 'En'), conditions=[f"CountryCode = '{country}'"])
    df_engines = filter_df_by_timestamp(df_all_engines, time)
    if engine_cat == '':
        other_engines = df_engines[df_engines['EngineCategory'].isna()]
        if other_engines.empty:
            DBOperations.instance.logger.info(f"No engines found for the given engine category {engine_cat}")
            return None
        other_engines['EngineType'] = ''
        return other_engines.groupby('EngineType').agg({'Code': list, 'CustomName': list, 'Performance': list, 'ID': list}).reset_index()
    elif engine_cat.lower() == 'all':
        df_engines['EngineType'] = ''
        return df_engines.groupby('EngineType').agg({'Code': list, 'CustomName': list, 'Performance': list, 'ID': list}).reset_index()
    
    df_engines = df_engines[df_engines['EngineCategory'].str.lower().str.strip() == engine_cat.lower().strip()]
    if df_engines.empty:
        DBOperations.instance.logger.info(f"No engines found for the given engine category {engine_cat}")
        return None
    
    df_engines['EngineType'] = df_engines['EngineType'].replace('', np.nan).fillna(df_engines['EngineCategory'])
    
    # group by the new column and return the list of engines for each group
    return df_engines.groupby('EngineType').agg({'Code': list, 'CustomName': list, 'Performance': list, 'ID': list}).reset_index()

def get_valid_pnos(country, model, time, engines_types):
    conditions=[f"CountryCode = '{country}'", f"Model = '{model}'", f"(Steering = '1' OR Steering = '1.0')"]
    allowed_engines = engines_types['Code'].explode('Code').unique().tolist()
    if not allowed_engines:
        DBOperations.instance.logger.info(f"No engines found for the given engine category {engines_types}")
    elif len(allowed_engines) == 1:
        conditions.append(f"Engine = '{allowed_engines[0]}'")
    else:
        conditions.append(f'Engine in {tuple(allowed_engines)}')

    df_all_pnos = DBOperations.instance.get_table_df(DBOperations.instance.config.get('AUTH', 'PNO'), conditions=conditions)
    df_pnos = filter_df_by_timestamp(df_all_pnos, time)
    if df_pnos.empty:
        DBOperations.instance.logger.info(f"No PNOs found for model {model}")
        raise Exception(f"No PNOs found for model {model}")

    return df_pnos
