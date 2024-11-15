from openpyxl import Workbook
from src.display_tables import features
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side
from io import BytesIO


def extract_features(country,
                     model,
                     model_year,
                     engine,
                     sales_version,
                     gearbox):
    df_pno_features = features.query_features(country = country,
                                              model = model,
                                              model_year = model_year,
                                              engine = engine,
                                              sales_version = sales_version,
                                              gearbox = gearbox)
    wb = Workbook()
    
    # Remove the default sheet created
    default_sheet = wb.active
    wb.remove(default_sheet)

    ws: Worksheet
    ws = wb.create_sheet(title='features')

    # df_pno_features.columns
    cols = ["Feature (Option)", "CPAM Text", "Market Text", "Feature Category"]
    ws.append(cols)
    for idx in range(len(cols)):
        cell = ws.cell(row=1, column=idx+1)
        cell.fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
        cell.font = Font(bold=True)
        if idx == 0:
            ws.column_dimensions[cell.column_letter].width = 20
        else:
            ws.column_dimensions[cell.column_letter].width = 50


    for row in df_pno_features.itertuples(index=False, name=None):
        ws.append(row)

    for row in ws[2:ws.max_row]:
        for cell in row[0:ws.max_column]:
            cell.alignment = Alignment(wrapText=True, vertical="center")
    
    ws.freeze_panes = 'A2'

    wb.active = wb.index(ws)
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
